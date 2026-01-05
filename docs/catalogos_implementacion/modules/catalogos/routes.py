from app.models.catalogos import * 
from app.modules.catalogos.forms import *
from app.models.users import SyUsuarios
from app.extensions import db  # Asegúrate de que db esté importado correctamente

import pandas as pd
import math
import os
import io
import re

from flask import Flask, Blueprint, render_template, flash, request, redirect, url_for, jsonify, session, make_response, send_file, g, abort
from datetime import datetime, timedelta, date
from pyreportjasper import PyReportJasper
from sqlalchemy import text, and_, func
from sqlalchemy.orm import aliased
from platform import python_version
from reportlab.lib.pagesizes import letter, A4, portrait, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph
from reportlab.lib import colors
from reportlab.lib.units import cm, inch  # Importar la unidad de centímetros
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from utils import datos_estructurados

import MySQLdb

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from flask_login import current_user
from urllib.parse import urlparse

catalogos_bp = Blueprint('catalogos', __name__)

'''@catalogos_bp.before_request
def proteger_rutas_catalogos():
    if not current_user.is_authenticated:
        return redirect(url_for('users.login'))
    urls_permitidas = session.get('urls_permitidas', [])
    if not any(request.path.startswith(url) for url in urls_permitidas):
        abort(403)  # Acceso denegado'''

@catalogos_bp.before_request
def proteger_rutas_administracion():
    if not current_user.is_authenticated:
        return redirect(url_for('users.login'))

    urls_nopermitidas = session.get('urls_nopermitidas', [])
    current_path = urlparse(request.path).path.lower()

    if any(current_path == urlparse(url).path.lower() for url in urls_nopermitidas):
        abort(403)

'''
@catalogos_bp.route('/cat_menus', methods=['GET', 'POST'])
def cat_menus():
    form = MenusForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_menu = form.txtId.data.upper()
            menu = form.txtMenu.data.upper()
            desc_menu = form.txtDescMenu.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_menus(:ve_opcion, :ve_id_menu, :ve_menu, :ve_desc_menu, :ve_id_usr, @vs_id_menu, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_menu': id_menu, 've_menu': menu, 've_desc_menu': desc_menu, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_menus'))

    menus = db.session.query(CatMenus).all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = MenusForm(data=form_data)
    else:
        form = MenusForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_menus.html', form=form, menus=menus)'''

@catalogos_bp.route('/cat_menus', methods=['GET', 'POST'])
def cat_menus():
    form = MenusForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_menu = form.txtId.data.upper()
            menu = form.txtMenu.data.upper()
            desc_menu = form.txtDescMenu.data.upper()
            clave_search = form.txtSearch.data.upper()  # <-- Nuevo campo
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text(
                    "CALL sp_menu_cat_menus(:ve_opcion, :ve_id_menu, :ve_menu, :ve_desc_menu, :ve_clave_search, :ve_id_usr, @vs_id_menu, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"
                ),
                {
                    've_opcion': opcion,
                    've_id_menu': id_menu,
                    've_menu': menu,
                    've_desc_menu': desc_menu,
                    've_clave_search': clave_search,  # <-- Pasamos el parámetro
                    've_id_usr': id_usr
                }
            )

            db.session.commit()

            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_menus'))

    menus = db.session.query(CatMenus).all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = MenusForm(data=form_data)
    else:
        form = MenusForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_menus.html', form=form, menus=menus)

@catalogos_bp.route('/cat_submenus', methods=['GET', 'POST'])
def cat_submenus():
    form = SubMenusForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_submenu = form.txtId.data.upper()
            id_menu = form.txtIdMenu.data
            print(f"ID_MENU recibido: {id_menu!r}")  # Para debug, verifica qué trae
            submenu = form.txtSubMenu.data.upper()
            desc_submenu = form.txtDescSubMenu.data.upper()
            url = form.txtUrl.data.upper()
            id_usr = "1".upper()


            result = db.session.execute(
                text("""CALL sp_menu_cat_submenus(
                    :ve_opcion, :ve_id_submenu, :ve_id_menu, :ve_submenu,
                    :ve_desc_submenu, :ve_url, :ve_id_usr,
                    @vs_id_submenu, @vs_resp, @vs_dsc_resp, @vs_bool_resp
                )"""),
                {
                    've_opcion': opcion,
                    've_id_submenu': id_submenu,
                    've_id_menu': id_menu,# ✅ Enviamos el ID correcto
                    've_submenu': submenu,
                    've_desc_submenu': desc_submenu,
                    've_url': url,
                    've_id_usr': id_usr
                }
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_submenus'))

    submenus = db.session.query(CatSubMenus,
                                CatMenus.menu).join(CatMenus, CatSubMenus.id_menu == CatMenus.id_menu).all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = SubMenusForm(data=form_data)
    else:
        form = SubMenusForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_submenus.html', form=form, submenus=submenus)

@catalogos_bp.route('/cat_clinicas', methods=['GET', 'POST'])
def cat_clinicas():
    print("Ruta cat_clinicas accedida")
    form = ClinicaForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()     
            id_clin = form.txtId.data.upper()
            clinica = form.txtClinica.data.upper()
            folio_clin = form.txtFolioClin.data.upper()
            interna = form.txtInterna.data.upper()
            ht_mat = form.txtHTMat.data.upper().strip()
            ht_ves = form.txtHTVes.data.upper().strip()
            ht_noc = form.txtHTNoc.data.upper().strip()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_clinicas(:ve_opcion, :ve_id_clin, :ve_clinica, :ve_folio_clin , :ve_interna, :ve_ht_mat, :ve_ht_ves, :ve_ht_noc, :ve_id_usr, @vs_id_clin, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_clin': id_clin, 've_clinica': clinica, 've_folio_clin': folio_clin, 've_interna': interna, 've_ht_mat': ht_mat, 've_ht_ves': ht_ves, 've_ht_noc': ht_noc, 've_id_usr': id_usr}
            )
            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            # result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_clin AS vs_id_clin"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_clinicas'))

    clinicas = CatClinicas.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = ClinicaForm(data=form_data)
    else:
        form = ClinicaForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_clinicas.html', form=form, clinicas=clinicas)

@catalogos_bp.route("/cat_especialidades", methods=['GET', 'POST'])
def cat_especialidades():
    form = EspecialidadesForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_espec = form.txtId.data
            especialidad = form.txtEspecialidad.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_especialidades(:ve_opcion, :ve_id_espec, :ve_especialidad, :ve_id_usr, @vs_id_espec, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_espec': id_espec.upper(), 've_especialidad': especialidad.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            # result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_especialidades'))

    especialidades = CatEspecialidades.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = EspecialidadesForm(data=form_data)
    else:
        form = EspecialidadesForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_especialidades.html', form=form, especialidades=especialidades)

@catalogos_bp.route('/cat_escuelas',methods=['GET', 'POST'])
def cat_escuelas():
    form=EscuelasForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_esc = form.txtId.data
            escuela = form.txtEscuela.data
            sigls_esc = form.txtSiglsEsc.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_escuelas(:ve_opcion, :ve_id_esc, :ve_escuela, :ve_sigls_esc, :ve_id_usr, @vs_id_esc, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_esc': id_esc.upper(), 've_escuela': escuela.upper(), 've_sigls_esc': sigls_esc.upper(), 've_id_usr': id_usr.upper()}
            )
            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_escuelas'))

    escuelas = CatEscuelas.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = EscuelasForm(data=form_data)
    else:
        form = EscuelasForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_escuelas.html', form=form, escuelas=escuelas)

@catalogos_bp.route('/cat_turnos', methods=['GET', 'POST'])
def cat_turnos():
    form = TurnosForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_trno = form.txtId.data
            turno = form.txtNombre.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_turnos(:ve_opcion, :ve_id_trno, :ve_turno, :ve_id_usr, @vs_id_trno, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_trno': id_trno.upper(), 've_turno': turno.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_turnos'))

    turnos = CatTurnos.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = TurnosForm(data=form_data)
    else:
        form = TurnosForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_turnos.html', form=form, turnos=turnos)

@catalogos_bp.route('/cat_enfermedades', methods=['GET', 'POST'])
def cat_enfermedades():
    form = EnfermedadesForm()
    
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            if opcion == '170' or opcion == '171':
                id_enf = ""
                cve_enf = ""
                enfermedad = ""
                vers_cie = form.txtCies.data
            else:
                id_enf = form.txtId.data
                cve_enf = form.txtCve.data
                enfermedad = form.txtEnfermedad.data
                vers_cie = form.txtVersCie.data
            id_usr = "1"


            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_enfermedades(:ve_opcion, :ve_id_enf, :ve_cve_enf, :ve_enfermedad, :ve_vers_cie, :ve_id_usr, @vs_id_enf, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_enf': id_enf.upper(), 've_cve_enf': cve_enf.upper(), 've_enfermedad': enfermedad.upper(), 've_vers_cie': vers_cie.upper(), 've_id_usr': id_usr.upper()}
            )
            
            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_enfermedades'))

    enfermedades = db.session.query(CatEnfermedades).all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = EnfermedadesForm(data=form_data)
    else:
        form = EnfermedadesForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_enfermedades.html', form=form, enfermedades=enfermedades)

# @catalogos_bp.route('/cat_enfermedades', methods=['GET', 'POST'])
# def cat_enfermedades():
#     form = EnfermedadesForm()
#     opciones_cies = CatEnfermedades.query.with_entities(CatEnfermedades.vers_cie, CatEnfermedades.est_enf).group_by(CatEnfermedades.vers_cie, CatEnfermedades.est_enf).all()
#     if request.method == 'POST':
#         try:
#             session['bool_data'] = "false"
#             opcion = form.opcion.data
#             id_enf = form.txtId.data
#             cve_enf = form.txtCve.data
#             enfermedad = form.txtEnfermedad.data
#             vers_cie = form.txtVersCie.data
#             id_usr = "1"

#             # Ejecutar el procedimiento almacenado
#             result = db.session.execute(
#                 text("CALL sp_menu_cat_enfermedades(:ve_opcion, :ve_id_enf, :ve_cve_enf, :ve_enfermedad, :ve_vers_cie, :ve_id_usr, @vs_id_enf, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
#                 {'ve_opcion': opcion.upper(), 've_id_enf': id_enf.upper(), 've_cve_enf': cve_enf.upper(), 've_enfermedad': enfermedad.upper(), 've_vers_cie': vers_cie.upper(), 've_id_usr': id_usr.upper()}
#             )

#             db.session.commit()

#             # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
#             #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_enf AS vs_id_enf"))
#             result_tuple = result.first()

#             vs_resp = result_tuple[-4]
#             vs_dsc_resp = result_tuple[-3]
#             vs_bool_resp = result_tuple[-2]

#             if vs_bool_resp == 'true':
#                 flash(vs_resp, 'success')
#             else:
#                 flash(vs_resp, 'error')
#                 session['bool_data'] = "true"
#                 session['form_data'] = request.form.to_dict(flat=True)
#                 session['form_data']['txtId'] = result_tuple[-1]

#             flash(vs_dsc_resp, 'mensaje')

#         except Exception as ex:
#             flash('Error al procesar el formulario', 'error')
#             flash(str(ex), 'error')
        
#         return redirect(url_for('cat_enfermedades'))

#     enfermedades = db.session.query(CatEnfermedades).all()

#     if session.get('bool_data') == "true":
#         form_data = session.get('form_data')
#         form = EnfermedadesForm(data=form_data)
#     else:
#         form = EnfermedadesForm()

#     session.pop('bool_data', None)
#     session.pop('form_data', None)

#     return render_template('cat/cat_enfermedades.html', form=form, enfermedades=enfermedades, opciones_cies=opciones_cies)

@catalogos_bp.route('/cat_escolaridad', methods=['GET', 'POST'])
def cat_escolaridad():
    form = EscolaridadForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_escol = form.txtId.data
            escolaridad = form.txtNombre.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_escolaridad(:ve_opcion, :ve_id_escol, :ve_escolaridad, :ve_id_usr, @vs_id_escol, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_escol': id_escol.upper(), 've_escolaridad': escolaridad.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_escol AS vs_id_escol"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_escolaridad'))

    escolaridades = CatEscolaridad.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = EscolaridadForm(data=form_data)
    else:
        form = EscolaridadForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_escolaridad.html', form=form, escolaridades=escolaridades)

@catalogos_bp.route('/cat_tipo_consulta', methods=['GET', 'POST'])
def cat_tipo_consulta():
    form = TipoConsultaForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_t_consul = form.txtId.data
            tipo_consulta = form.txtNombre.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_tipo_consulta(:ve_opcion, :ve_id_t_consul, :ve_tipo_consulta, :ve_id_usr, @vs_id_t_consul, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_t_consul': id_t_consul.upper(), 've_tipo_consulta': tipo_consulta.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_t_consul AS vs_id_t_consul"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_tipo_consulta'))

    tipos_consulta = CatTipoConsulta.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = TipoConsultaForm(data=form_data)
    else:
        form = TipoConsultaForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_tipo_consulta.html', form=form, tipos_consulta=tipos_consulta)

@catalogos_bp.route('/cat_cie10', methods=['GET', 'POST'])
def cat_cie10():
    form = CIE10Form()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_cie10 = form.txtId.data.upper()
            cve_cie10 = form.txtCve.data.upper()
            desc_cie10 = form.txtDesc.data.upper()
            sexo = form.txtSexo.data.upper()
            id_espec = form.txtIdEspec.data.upper()
            desc_espec = form.txtDescEspec.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_cie10(:ve_opcion, :ve_id_cie10, :ve_cve_cie10, :ve_desc_cie10, :ve_sexo, :ve_id_espec, :ve_desc_espec, :ve_id_usr, @vs_id_cie10, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_cie10': id_cie10, 've_cve_cie10': cve_cie10, 've_desc_cie10': desc_cie10, 've_sexo': sexo, 've_id_espec': id_espec, 've_desc_espec': desc_espec, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_cie10 AS vs_id_cie10"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_cie10'))

    cies = db.session.query(CatCIE10, CatEspecialidades.especialidad).join(CatEspecialidades, CatCIE10.id_espec == CatEspecialidades.id_espec).all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = CIE10Form(data=form_data)
    else:
        form = CIE10Form()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_cie10.html', form=form, cies=cies)

@catalogos_bp.route('/cat_cie11', methods=['GET', 'POST'])
def cat_cie11():
    form = CIE11Form()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_cie11 = form.txtId.data.upper()
            cve_cie11 = form.txtCve.data.upper()
            desc_cie11 = form.txtDesc.data.upper()
            sexo = form.txtSexo.data.upper()
            id_espec = form.txtIdEspec.data.upper()
            desc_espec = form.txtDescEspec.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_cie11(:ve_opcion, :ve_id_cie11, :ve_cve_cie11, :ve_desc_cie11, :ve_sexo, :ve_id_espec, :ve_desc_espec, :ve_id_usr, @vs_id_cie11, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_cie11': id_cie11, 've_cve_cie11': cve_cie11, 've_desc_cie11': desc_cie11, 've_sexo': sexo, 've_id_espec': id_espec, 've_desc_espec': desc_espec, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_cie11 AS vs_id_cie11"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_cie11'))

    cies = db.session.query(CatCIE11, CatEspecialidades.especialidad).join(CatEspecialidades, CatCIE11.id_espec == CatEspecialidades.id_espec).all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = CIE11Form(data=form_data)
    else:
        form = CIE11Form()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_cie11.html', form=form, cies=cies)

@catalogos_bp.route('/cat_tipo_hospn', methods=['GET', 'POST'])
def cat_tipo_hospn():
    form = TipoHospnForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_t_hospn = form.txtId.data
            tipo_hospn = form.txtNombre.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_tipo_hospn(:ve_opcion, :ve_id_t_hospn, :ve_tipo_hospn, :ve_id_usr, @vs_id_t_hospn, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_t_hospn': id_t_hospn.upper(), 've_tipo_hospn': tipo_hospn.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_t_hospn AS vs_id_t_hospn"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_tipo_hospn'))

    tipos_hospn = CatTipoHospn.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = TipoHospnForm(data=form_data)
    else:
        form = TipoHospnForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_tipo_hospn.html', form=form, tipos_hospn=tipos_hospn)

@catalogos_bp.route('/cat_calidadlab', methods=['GET', 'POST'])
def cat_calidadlab():
    form = CalidadLabForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_calidadlab = form.txtCve.data
            calidadlab = form.txtCalidadLab.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_calidadlab(:ve_opcion, :ve_id_calidadlab, :ve_calidadlab, :ve_id_usr, @vs_id_calidadlab, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_calidadlab': id_calidadlab.upper(), 've_calidadlab': calidadlab.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_calidadlab AS vs_id_calidadlab"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtCve'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_calidadlab'))

    datos = CatCalidadLab.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = CalidadLabForm(data=form_data)
    else:
        form = CalidadLabForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_calidadlab.html', form=form, datos=datos)

@catalogos_bp.route('/cat_edocivil', methods=['GET', 'POST'])
def cat_edocivil():
    form = EdoCivilForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_edocivil = form.txtId.data
            edocivil = form.txtEdoCivil.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_edocivil(:ve_opcion, :ve_id_edocivil, :ve_edocivil, :ve_id_usr, @vs_id_edocivil, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_edocivil': id_edocivil.upper(), 've_edocivil': edocivil.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_edocivil AS vs_id_edocivil"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_edocivil'))

    datos = CatEdoCivil.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = EdoCivilForm(data=form_data)
    else:
        form = EdoCivilForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_edocivil.html', form=form, datos=datos)

@catalogos_bp.route('/cat_tpsanguineo', methods=['GET', 'POST'])
def cat_tpsanguineo():
    form = TpSanguineoForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_tpsanguineo = form.txtId.data
            tpsanguineo = form.txtTpSanguineo.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_tpsanguineo(:ve_opcion, :ve_id_tpsanguineo, :ve_tpsanguineo, :ve_id_usr, @vs_id_tpsanguineo, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_tpsanguineo': id_tpsanguineo.upper(), 've_tpsanguineo': tpsanguineo.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_tpsanguineo AS vs_id_tpsanguineo"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_tpsanguineo'))

    datos = CatTpSanguineo.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = TpSanguineoForm(data=form_data)
    else:
        form = TpSanguineoForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_tpsanguineo.html', form=form, datos=datos)

@catalogos_bp.route('/cat_gpomedic', methods=['GET', 'POST'])
def cat_gpomedic():
    form = GpoMedicForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_gpomedic = form.txtId.data.upper()
            gpomedic = form.txtGpoMedic.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_gpomedic(:ve_opcion, :ve_id_gpomedic, :ve_gpomedic, :ve_id_usr, @vs_id_gpomedic, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_gpomedic': id_gpomedic, 've_gpomedic': gpomedic, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_gpomedic'))

    gposmedic = CatGpoMedic.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = GpoMedicForm(data=form_data)
    else:
        form = GpoMedicForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_gpomedic.html', form=form, gposmedic=gposmedic)

@catalogos_bp.route('/cat_hospitales', methods=['GET', 'POST'])
def cat_hospitales():
    form = HospitalesForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_hosp = form.txtId.data.upper()
            hospital = form.txtHospital.data.upper()
            calle = form.txtCalle.data.upper()
            asenta = form.txtAsenta.data.upper()
            cp = form.txtCp.data.upper()
            municipio = form.txtMunicipio.data.upper()
            entidad_fed = form.txtEntidadFed.data.upper()
            telefono = form.txtTelefono.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_hospitales(:ve_opcion, :ve_id_hosp, :ve_hospital, :ve_calle, :ve_asenta, :ve_cp, :ve_municipio, :ve_entidad_fed, :ve_telefono, :ve_id_usr, @vs_id_hosp, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_hosp': id_hosp, 've_hospital': hospital, 've_calle': calle, 've_asenta': asenta, 've_cp': cp, 've_municipio': municipio, 've_entidad_fed': entidad_fed, 've_telefono': telefono, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_hospitales'))

    hospitales = CatHospitales.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = HospitalesForm(data=form_data)
    else:
        form = HospitalesForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_hospitales.html', form=form, hospitales=hospitales)

@catalogos_bp.route('/cat_origencons', methods=['GET', 'POST'])
def cat_origencons():
    form = OrigenConsForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_origencons = form.txtCve.data
            origencons = form.txtOrigenCons.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_origencons(:ve_opcion, :ve_id_origencons, :ve_origencons, :ve_id_usr, @vs_id_origencons, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_origencons': id_origencons.upper(), 've_origencons': origencons.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_calidadlab AS vs_id_calidadlab"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtCve'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_origencons'))

    consulta = CatOrigenCons.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = OrigenConsForm(data=form_data)
    else:
        form = OrigenConsForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_origencons.html', form=form, consulta=consulta)

@catalogos_bp.route('/cat_ocupaciones', methods=['GET', 'POST'])
def cat_ocupaciones():
    form = OcupacionesForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_ocupacion = form.txtId.data
            ocupacion = form.txtOcupacion.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_ocupaciones(:ve_opcion, :ve_id_ocupacion, :ve_ocupacion, :ve_id_usr, @vs_id_ocupacion, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_ocupacion': id_ocupacion.upper(), 've_ocupacion': ocupacion.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_ocupaciones'))

    ocupaciones = CatOcupaciones.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = OcupacionesForm(data=form_data)
    else:
        form = OcupacionesForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_ocupaciones.html', form=form, ocupaciones=ocupaciones)

@catalogos_bp.route('/cat_parentescos', methods=['GET', 'POST'])
def cat_parentescos():
    form = ParentescosForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_parentesco = form.txtCve.data
            parentesco = form.txtParentesco.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_parentescos(:ve_opcion, :ve_id_parentesco, :ve_parentesco, :ve_id_usr, @vs_id_parentesco, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_parentesco': id_parentesco.upper(), 've_parentesco': parentesco.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_calidadlab AS vs_id_calidadlab"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtCve'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_parentescos'))

    parentesco = CatParentescos.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = ParentescosForm(data=form_data)
    else:
        form = ParentescosForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_parentescos.html', form=form, parentesco=parentesco)

@catalogos_bp.route('/cat_laboratorios', methods=['GET', 'POST'])
def cat_laboratorios():
    form = LaboratoriosForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_lab = form.txtId.data.upper()
            laboratorio = form.txtLaboratorio.data.upper()
            unidad_lab = form.txtUnidadLab.data.upper()
            calle = form.txtCalle.data
            asenta = form.txtAsenta.data.upper()
            cp = form.txtCp.data.upper()
            municipio = form.txtMunicipio.data.upper()
            entidad_fed = form.txtEntidadFed.data.upper()
            telefono = form.txtTelefono.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_laboratorios(:ve_opcion, :ve_id_lab, :ve_laboratorio, :ve_unidad_lab, :ve_calle, :ve_asenta, :ve_cp, :ve_municipio, :ve_entidad_fed, :ve_telefono, :ve_id_usr, @vs_id_hosp, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_lab': id_lab, 've_laboratorio': laboratorio, 've_unidad_lab': unidad_lab, 've_calle': calle, 've_asenta': asenta, 've_cp': cp, 've_municipio': municipio, 've_entidad_fed': entidad_fed, 've_telefono': telefono, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_laboratorios'))

    laboratorios = CatLaboratorios.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = LaboratoriosForm(data=form_data)
    else:
        form = LaboratoriosForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_laboratorios.html', form=form, laboratorios=laboratorios)

@catalogos_bp.route('/cat_tplicencia', methods=['GET', 'POST'])
def cat_tplicencia():
    form = TpLicenciaForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_tplicencia = form.txtId.data.upper()
            tplicencia = form.txtTpLicencia.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_tplicencia(:ve_opcion, :ve_id_tplicencia, :ve_tplicencia, :ve_id_usr, @vs_id_tplicencia, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_tplicencia': id_tplicencia, 've_tplicencia': tplicencia, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_tplicencia'))

    licencia = CatTipoLicencia.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = TpLicenciaForm(data=form_data)
    else:
        form = TpLicenciaForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_tplicencia.html', form=form, licencia=licencia)

@catalogos_bp.route('/cat_tpautorizacion', methods=['GET', 'POST'])
def cat_tpautorizacion():
    form = TpAutorizacionForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_tpautorizacion = form.txtId.data
            cve_tpautorizacion = form.txtCve.data
            tpautorizacion = form.txtTpAutorizacion.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_tpautorizacion(:ve_opcion, :ve_id_tpautorizacion, :ve_cve_tpautorizacion, :ve_tpautorizacion, :ve_id_usr, @vs_id_t_hospn, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_tpautorizacion': id_tpautorizacion.upper(), 've_cve_tpautorizacion': cve_tpautorizacion.upper(), 've_tpautorizacion': tpautorizacion.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_t_hospn AS vs_id_t_hospn"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_tpautorizacion'))

    tpautorizacion = CatTpAutorizacion.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = TpAutorizacionForm(data=form_data)
    else:
        form = TpAutorizacionForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_tpautorizacion.html', form=form, tpautorizacion=tpautorizacion)

@catalogos_bp.route('/cat_tpcitas', methods=['GET', 'POST'])
def cat_tpcitas():
    form = TpCitasForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_tpcita = form.txtId.data.upper()
            tpcita = form.txtTpCita.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_tpcitas(:ve_opcion, :ve_id_tpcita, :ve_tpcita, :ve_id_usr, @vs_id_, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_tpcita': id_tpcita, 've_tpcita': tpcita, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_tpcitas'))

    datos = CatTpCitas.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = TpCitasForm(data=form_data)
    else:
        form = TpCitasForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_tpcitas.html', form=form, datos=datos)

@catalogos_bp.route('/cat_pases', methods=['GET', 'POST'])
def cat_pases():
    form = PasesForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_pase = form.txtId.data.upper()
            pase = form.txtPase.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_pases(:ve_opcion, :ve_id_pase, :ve_pase, :ve_id_usr, @vs_id_, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_pase': id_pase, 've_pase': pase, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_pases'))

    pases = CatPases.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = PasesForm(data=form_data)
    else:
        form = PasesForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_pases.html', form=form, pases=pases)

@catalogos_bp.route('/cat_serviciosclin', methods=['GET', 'POST'])
def cat_serviciosclin():
    form = ServiciosClinForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_serviciosclin = form.txtId.data.upper()
            serviciosclin = form.txtServiciosClin.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_serviciosclin(:ve_opcion, :ve_id_serviciosclin, :ve_serviciosclin, :ve_id_usr, @vs_id_, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_serviciosclin': id_serviciosclin, 've_serviciosclin': serviciosclin, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_serviciosclin'))

    datos = CatServiciosClin.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = ServiciosClinForm(data=form_data)
    else:
        form = ServiciosClinForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_serviciosclin.html', form=form, datos=datos)

@catalogos_bp.route('/cat_estudiosmed', methods=['GET', 'POST'])
def cat_estudiosmed():
    form = EstudiosMedForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_estudiomed = form.txtId.data.upper()
            estudiomed = form.txtEstudioMed.data.upper()
            valor = form.txtValor.data
            tp_estudiomed = form.txtTpEstudioMed.data.upper()
            indicacion = form.txtIndicacion.data.upper()
            estudiogral = form.txtEstudioGral.data.upper()
            autorizado = form.txtAutorizado.data.upper()
            tp_grupo = form.txtTpGrupo.data.upper()
            id_provedor = form.txtIdProvedor.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_estudiosmed(:ve_opcion, :ve_id_estudiomed, :ve_estudiomed, :ve_valor, :ve_tp_estudiomed, :ve_indicacion, :ve_estudiogral, :ve_autorizado, :ve_tp_grupo, :ve_id_provedor, :ve_id_usr, @vs_id_estudiomed, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_estudiomed': id_estudiomed, 've_estudiomed': estudiomed, 've_valor': valor, 've_tp_estudiomed': tp_estudiomed, 've_indicacion': indicacion, 've_estudiogral': estudiogral, 've_autorizado': autorizado, 've_tp_grupo': tp_grupo, 've_id_provedor': id_provedor, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_estudiosmed'))

    estudios = CatEstudiosMed.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = EstudiosMedForm(data=form_data)
    else:
        form = EstudiosMedForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_estudiosmed.html', form=form, estudios=estudios)

@catalogos_bp.route('/cat_consultorios', methods=['GET', 'POST'])
def cat_consultorios():
    form = ConsultoriosForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_consult = form.txtId.data.upper()
            no_consult = form.txtNoConsult.data.upper()
            turno = form.txtIdTrno.data.upper()
            clinica = form.txtIdClin.data.upper()
            consult = form.txtConsult.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_consultorios(:ve_opcion, :ve_id_consult, :ve_no_consult, :ve_turno, :ve_clinica, :ve_consult, :ve_id_usr, @vs_id_consult, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_consult': id_consult, 've_no_consult': no_consult, 've_turno': turno, 've_clinica': clinica, 've_consult': consult, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_consultorios'))

    consultorios = db.session.query(CatConsultorios).all()
    
    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = ConsultoriosForm(data=form_data)
    else:
        form = ConsultoriosForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_consultorios.html', form=form, consultorios=consultorios)

@catalogos_bp.route('/cat_tpareas', methods=['GET', 'POST'])
def cat_tpareas():
    form = TpAreasForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_tparea = form.txtId.data
            tparea = form.txtTpArea.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_tpareas(:ve_opcion, :ve_id_tparea, :ve_tparea, :ve_id_usr, @vs_id_tparea, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_tparea': id_tparea.upper(), 've_tparea': tparea.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_edocivil AS vs_id_edocivil"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_tpareas'))

    tparea = CatTpAreas.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = TpAreasForm(data=form_data)
    else:
        form = TpAreasForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_tpareas.html', form=form, tparea=tparea)

@catalogos_bp.route('/cat_areas', methods=['GET', 'POST'])
def cat_areas():
    form = AreasForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_area = form.txtId.data.upper()
            area = form.txtArea.data.upper()
            tparea = form.txtIdTpArea.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_areas(:ve_opcion, :ve_id_area, :ve_area, :ve_tparea, :ve_id_usr, @vs_id_area, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_area': id_area, 've_area': area, 've_tparea': tparea, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_areas'))

    areas = db.session.query(CatAreas).all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = AreasForm(data=form_data)
    else:
        form = AreasForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_areas.html', form=form, areas=areas)

@catalogos_bp.route('/cat_tpbajas', methods=['GET', 'POST'])
def cat_tpbajas():
    form = TpBajasForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_tpbaja = form.txtCve.data
            tpbaja = form.txtTpBaja.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_tpbajas(:ve_opcion, :ve_id_tpbaja, :ve_tpbaja, :ve_id_usr, @vs_id_tpbaja, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_tpbaja': id_tpbaja.upper(), 've_tpbaja': tpbaja.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_calidadlab AS vs_id_calidadlab"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtCve'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_tpbajas'))

    baja = CatTpBajas.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = TpBajasForm(data=form_data)
    else:
        form = TpBajasForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_tpbajas.html', form=form, baja=baja)

@catalogos_bp.route('/cat_cies', methods=['GET', 'POST'])
def cat_cies():
    form = CiesForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            global df_global 
            global total_errores
            total_datos = len(df_global)
            opcion = "153"
            id_enf = ""
            id_usr = "1".upper()

            for index, row in df_global.iterrows():
                if row['TIPO DE ERROR'] == "":
                    cve_enf = row[0].upper()
                    enfermedad = row[1].upper()
                    vers_cie = row[2].upper()

                    result = db.session.execute(
                        text("CALL sp_menu_cat_enfermedades(:ve_opcion, :ve_id_enf, :ve_cve_enf, :ve_enfermedad, :ve_vers_cie, :ve_id_usr, @vs_id_enf, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                        {'ve_opcion': opcion, 've_id_enf': id_enf, 've_cve_enf': cve_enf, 've_enfermedad': enfermedad, 've_vers_cie': vers_cie, 've_id_usr': id_usr}
                    )

                    db.session.commit()

                    result_tuple = result.first()
                    vs_resp = result_tuple[-4]
                    vs_dsc_resp = result_tuple[-3]
                    vs_bool_resp = result_tuple[-2]

                    if vs_bool_resp == 'false':
                        row[3] = vs_dsc_resp;
    
            
            df_errores = df_global[df_global['TIPO DE ERROR'] != ""]
            total_errores = len(df_errores)
            total_cargados = total_datos - total_errores

            session['bool_data'] = 'true'
            session['total_errores'] = total_errores
            session['total_cargados'] = total_cargados
            session['total_datos'] = total_datos

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_cies'))

    if session.get('bool_data') == "true":
        total_errores = session.get('total_errores')
        total_cargados = session.get('total_cargados')
        total_datos = session.get('total_datos')

        session.pop('bool_data', None)
        session.pop('total_errores', None)
        session.pop('total_cargados', None)
        session.pop('total_datos', None)
        
        return render_template('catalogos/cat_cies.html', form=form, datos_cargados='True', total_errores=total_errores, total_cargados=total_cargados, total_datos=total_datos)
    else:
        form = CiesForm()
        return render_template('catalogos/cat_cies.html', form=form)
    
@catalogos_bp.route('/cat_tpautorizador', methods=['GET', 'POST'])
def cat_tpautorizador():
    form = TpAutorizadorForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data
            id_tpautorizador = form.txtCve.data
            tpautorizador = form.txtTpAutorizador.data
            id_usr = "1"

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_tpautorizador(:ve_opcion, :ve_id_tpautorizador, :ve_tpautorizador, :ve_id_usr, @vs_id_tpautorizador, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion.upper(), 've_id_tpautorizador': id_tpautorizador.upper(), 've_tpautorizador': tpautorizador.upper(), 've_id_usr': id_usr.upper()}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_calidadlab AS vs_id_tpautorizado"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtCve'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_tpautorizador'))

    autorizador = CatTpAutorizador.query.all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = TpAutorizadorForm(data=form_data)
    else:
        form = TpAutorizadorForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_tpautorizador.html', form=form, autorizador=autorizador)    

@catalogos_bp.route('/cat_autorizadores', methods=['GET', 'POST'])
def cat_autorizadores():
    form = AutorizadoresForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_autorizador = form.txtId.data.upper()
            id_clin = form.txtCve.data.upper()
            clinica = form.txtClinica.data.upper()
            autorizador = form.txtAutorizador.data.upper()
            cargo = form.txtCargo.data.upper()
            id_tpautorizador = form.txtIdTpAutorizador
            tpautorizador = form.txtTpAutorizador.data.upper()
            img_firma = form.txtImgFirma.data.upper()
            pwd_autorizador = form.txtPwdAutorizador.data.upper()
            id_usuario = form.txtIdUser.data.upper()
            usuario = form.txtUsuario.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_autorizadores(:ve_opcion, :ve_id_autorizador, :ve_id_clin, :ve_clinica, :ve_autorizador, :ve_cargo, :ve_id_tpautorizado :ve_tpautorizado, :ve_img_firma, :ve_pwd_autz, ve_id_usuario, ve_expediente, :ve_usr,  :ve_id_usr, @vs_id_autorizador, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_autorizador': id_autorizador, 've_id_clin': id_clin, 've_clinica': clinica, 've_autorizador': autorizador, 've_cargo': cargo, 've_id_tpautorizador': id_tpautorizador, 've_tpautorizador': tpautorizador, 've_img_firma': img_firma, 've_pwd_autorizador': pwd_autorizador, 've_id_usuario': id_usuario, 've_usuario': usuario, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_autorizadores'))

    autorizadores = db.session.query(
        CatAutorizadores,
        CatTpAutorizador.tpautorizador, 
        CatClinicas.clinica,
        SyUsuarios.expediente,
        SyUsuarios.nombre,
        SyUsuarios.paterno,
        SyUsuarios.materno
    ).join(
        CatClinicas, CatAutorizadores.id_clin == CatClinicas.id_clin
    ).join(
        SyUsuarios, CatAutorizadores.id_usuario == SyUsuarios.id_usuario
    ).join(
        CatTpAutorizador, CatAutorizadores.id_tpautorizador == CatTpAutorizador.id_tpautorizador
    ).all()
    
    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = AutorizadoresForm(data=form_data)
    else:
        form = AutorizadoresForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_autorizadores.html', form=form, autorizadores=autorizadores)

@catalogos_bp.route('/cat_medicosclin', methods=['GET', 'POST'])
def cat_medicosclin():
    form = MedicosClinForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_medclin = form.txtId.data.upper()
            usuario = form.txtUsuario.data.upper()
            sexo = form.txtSexo.data.upper()
            fch_nac = form.txtFchNac.data.upper()
            cedula = form.txtCedula.data.upper()
            escuela = form.txtEscuela.data.upper()
            especialidad = form.txtEspecialidad.data.upper()
            serviciosclin = form.txtServiciosClin.data.upper()
            clinica = form.txtClinica.data.upper()
            consult = form.txtConsultorio.data.upper()
            hr_ini = form.txtHrIni.data.upper().strip()
            hr_term = form.txtHrTerm.data.upper().strip()
            interv_consul = form.txtIntervConsul.data.upper()
            direccion = form.txtDireccion.data.upper()
            dias = form.txtDias.data.upper()
            ambos_turn = form.txtAmbosTurnos.data.upper()
            consult2 = form.txtConsultorio2.data.upper()
            hr_ini2 = form.txtHrInic2.data.upper().strip()
            hr_term2 = form.txtHrTerm2.data.upper().strip()
            interv_consul2 = form.txtIntervConsul2.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_medicosclin(:ve_opcion, :ve_id_medclin, :ve_usuario, :ve_sexo, :ve_fch_nac, :ve_cedula, :ve_escuela, :ve_especialidad, :ve_serviciosclin, :ve_clinica, :ve_consult, :ve_hr_ini, :ve_hr_term, :ve_interv_consul, :ve_direccion, :ve_dias, :ve_ambos_turn,  :ve_consult2, :ve_hr_ini2, :ve_hr_term2, :ve_interv_consul2, :ve_id_usr, @vs_id_medclin, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_medclin': id_medclin, 've_usuario': usuario, 've_sexo': sexo, 've_fch_nac': fch_nac, 've_cedula': cedula, 've_escuela': escuela, 've_especialidad': especialidad, 've_serviciosclin': serviciosclin, 've_clinica': clinica, 've_consult': consult, 've_hr_ini': hr_ini, 've_hr_term': hr_term, 've_interv_consul': interv_consul, 've_direccion': direccion, 've_dias': dias, 've_ambos_turn': ambos_turn, 've_consult2': consult2, 've_hr_ini2': hr_ini2, 've_hr_term2': hr_term2, 've_interv_consul2': interv_consul2, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_medclin"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_medicosclin'))

    # Crear alias para el segundo consultorio
    Consultorio1 = aliased(CatConsultorios)
    Consultorio2 = aliased(CatConsultorios)

    medicsclin = db.session.query(
        CatMedicosClin,
        CatEscuelas.sigls_esc,
        CatEscuelas.escuela,
        CatEspecialidades.especialidad,
        CatServiciosClin.serviciosclin,
        CatClinicas.clinica,
        Consultorio1.no_consult.label('no_consult1'),
        Consultorio1.turno.label('turno1'),
        Consultorio1.consult.label('consultorio1'),
        Consultorio2.no_consult.label('no_consult2'),
        Consultorio2.turno.label('turno2'),
        Consultorio2.consult.label('consultorio2'),
        SyUsuarios.expediente,
        SyUsuarios.nombre,
        SyUsuarios.paterno,
        SyUsuarios.materno
    ).join(
        CatEscuelas, CatMedicosClin.id_esc == CatEscuelas.id_esc
    ).join(
        CatEspecialidades, CatMedicosClin.id_espec == CatEspecialidades.id_espec
    ).join(
        CatServiciosClin, CatMedicosClin.id_serviciosclin == CatServiciosClin.id_serviciosclin
    ).join(
        Consultorio1, CatMedicosClin.id_consult == Consultorio1.id_consult
    ).outerjoin(
        Consultorio2, CatMedicosClin.id_consult2 == Consultorio2.id_consult
    ).join(
        CatClinicas, CatMedicosClin.id_clin == CatClinicas.id_clin
    ).join(
        SyUsuarios, CatMedicosClin.id_usuario == SyUsuarios.id_usuario
    ).all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = MedicosClinForm(data=form_data)
    else:
        form = MedicosClinForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_medicosclin.html', form=form, medicsclin=medicsclin)

@catalogos_bp.route('/cat_medicosespechosp', methods=['GET', 'POST'])
def cat_medicosespechosp():
    form = MedicosEspecHospForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_medespec = form.txtId.data.upper()
            id_usuario = form.txtIdUser.data.upper()
            usermedespec = form.txtCve.data.upper()
            usuario = form.txtUsuario.data.upper()
            tp_med = form.txtTpMed.data.upper()
            sexo = form.txtSexo.data.upper()
            fch_nac = form.txtFchNac.data.upper()
            cedula = form.txtCedula.data.upper()
            id_esc = form.txtIdEscuela.data.upper()
            escuela = form.txtEscuela.data.upper()
            id_espec = form.txtIdEspecialidad.data.upper()
            especialidad = form.txtEspecialidad.data.upper()
            id_serviciosclin = form.txtIdServiciosClin.data.upper()
            serviciosclin = form.txtServiciosClin.data.upper()
            id_clin = form.txtIdClinica.data.upper()
            clinica = form.txtClinica.data.upper()
            telefono = form.txtTelefono.data.upper()
            direccion = form.txtDireccion.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_medicosespechosp(:ve_opcion, :ve_id_medespec, :ve_id_usuario, :ve_usermedespec, :ve_usuario, :ve_sexo, :ve_fch_nac, :ve_cedula, :ve_id_esc, :ve_escuela, :ve_id_espec, :ve_especialidad, :ve_id_serviciosclin, :ve_serviciosclin, :ve_id_clin, :ve_clinica, :ve_telefono, :ve_direccion, :ve_id_usr, @vs_id_medespec, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_medespec': id_medespec, 've_id_usuario': id_usuario, 've_usermedespec': usermedespec, 've_usuario': usuario, 've_sexo': sexo, 've_fch_nac': fch_nac, 've_cedula': cedula, 've_id_esc': id_esc, 've_escuela': escuela, 've_id_espec': id_espec, 've_especialidad': especialidad, 've_id_serviciosclin': id_serviciosclin, 've_serviciosclin': serviciosclin, 've_id_clin': id_clin, 've_clinica': clinica, 've_telefono': telefono, 've_direccion': direccion, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_medclin"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_medicosespechosp'))

    medicsespec = db.session.query(
        CatMedicosEspecHosp,
        #CatRoles.rol,
        #CatRoles.tp_rol,
        CatEscuelas.escuela,
        CatEscuelas.sigls_esc,
        CatEspecialidades.especialidad,
        CatServiciosClin.serviciosclin,
        CatClinicas.folio_clin, 
        CatClinicas.clinica,
        SyUsuarios.expediente,
        SyUsuarios.nombre,
        SyUsuarios.paterno,
        SyUsuarios.materno
    #).join(
    #    CatRoles, CatMedicosEspecHosp.id_rol == CatRoles.id_rol
    ).join(
        CatEscuelas, CatMedicosEspecHosp.id_esc == CatEscuelas.id_esc
    ).join(
        CatEspecialidades, CatMedicosEspecHosp.id_espec == CatEspecialidades.id_espec
    ).join(
        CatServiciosClin, CatMedicosEspecHosp.id_serviciosclin == CatServiciosClin.id_serviciosclin
    ).join(
        CatClinicas, CatMedicosEspecHosp.id_clin == CatClinicas.id_clin
    ).join(
        SyUsuarios, CatMedicosEspecHosp.id_usuario == SyUsuarios.id_usuario
    ).all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = MedicosEspecHospForm(data=form_data)
    else:
        form = MedicosEspecHospForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_medicosespechosp.html', form=form, medicsespec=medicsespec)

@catalogos_bp.route('/cat_roles', methods=['GET', 'POST'])
def cat_roles():
    form = RolesForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_rol = form.txtId.data.upper()
            rol = form.txtRol.data.upper()
            tp_rol = form.txtTpRol.data.upper()
            desc_rol = form.txtDescRol.data.upper()
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_roles(:ve_opcion, :ve_id_rol, :ve_rol, :ve_tp_rol, :ve_desc_rol, :ve_id_usr, @vs_id_rol, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_rol': id_rol, 've_rol': rol, 've_tp_rol': tp_rol, 've_desc_rol': desc_rol, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_roles'))

    roles = db.session.query(CatRoles).all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = RolesForm(data=form_data)
    else:
        form = RolesForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_roles.html', form=form, roles=roles)


@catalogos_bp.route('/cat_medicamentos', methods=['GET', 'POST'])
def cat_medicamentos():
    form = MedicamentosForm()
    if request.method == 'POST':
        try:
            session['bool_data'] = "false"
            opcion = form.opcion.data.upper()
            id_medicamento = form.txtId.data.upper()
            cve = form.txtCve.data.upper()
            medicamento = form.txtMedicamento.data.upper()
            desc_medicamento = form.txtDescMedicamento.data.upper()
            med_disponible = form.txtMedDisponible.data
            surtido = form.txtSurtido.data
            id_usr = "1".upper()

            # Ejecutar el procedimiento almacenado
            result = db.session.execute(
                text("CALL sp_menu_cat_medicamentos(:ve_opcion, :ve_id_medicamento, :ve_cve, :ve_medicamento, :ve_desc_medicamento, :ve_med_disponible, :ve_surtido, :ve_id_usr, @vs_id_medicamento, @vs_resp, @vs_dsc_resp, @vs_bool_resp)"),
                {'ve_opcion': opcion, 've_id_medicamento': id_medicamento, 've_cve': cve, 've_medicamento': medicamento, 've_desc_medicamento': desc_medicamento, 've_med_disponible': med_disponible, 've_surtido': surtido, 've_id_usr': id_usr}
            )

            db.session.commit()

            # Recuperar los resultados de los parámetros de salida mediante una segunda consulta
            #result = db.session.execute(text("SELECT @vs_resp AS vs_resp, @vs_dsc_resp AS vs_dsc_resp, @vs_bool_resp AS vs_bool_resp, @vs_id_trno AS vs_id_trno"))
            result_tuple = result.first()

            vs_resp = result_tuple[-4]
            vs_dsc_resp = result_tuple[-3]
            vs_bool_resp = result_tuple[-2]

            if vs_bool_resp == 'true':
                flash(vs_resp, 'success')
            else:
                flash(vs_resp, 'error')
                session['bool_data'] = "true"
                session['form_data'] = request.form.to_dict(flat=True)
                session['form_data']['txtId'] = result_tuple[-1]

            flash(vs_dsc_resp, 'mensaje')

        except Exception as ex:
            flash('Error al procesar el formulario', 'error')
            flash(str(ex), 'error')
        
        return redirect(url_for('catalogos.cat_medicamentos'))

    medicamentos = db.session.query(CatMedicamentos).all()

    if session.get('bool_data') == "true":
        form_data = session.get('form_data')
        form = MedicamentosForm(data=form_data)
    else:
        form = MedicamentosForm()

    session.pop('bool_data', None)
    session.pop('form_data', None)

    return render_template('catalogos/cat_medicamentos.html', form=form, medicamentos=medicamentos)
####################################################################################################################################################################

@catalogos_bp.route('/actualizar_especialidades', methods=['GET'])
def actualizar_especialidades():
    especialidades = [(espec.id_espec, espec.especialidad) for espec in CatEspecialidades.query.filter_by(est_espec='A').order_by(CatEspecialidades.especialidad)]
    return jsonify(especialidades)

@catalogos_bp.route('/actualizar_tpareas', methods=['GET'])
def actualizar_tpareas():
    tpareas = [('', '')] + [(tparea.tparea, tparea.tparea) for tparea in CatTpAreas.query.filter_by(est_tparea='A').order_by(CatTpAreas.tparea)]
    return jsonify(tpareas)

@catalogos_bp.route('/actualizar_turnos', methods=['GET'])
def actualizar_turnos():
    turnos = [('', '')] + [(trno.turno, trno.turno) for trno in CatTurnos.query.filter_by(est_trno='A').order_by(CatTurnos.turno)]
    return jsonify(turnos)

@catalogos_bp.route('/actualizar_clinicas', methods=['GET'])
def actualizar_clinicas():
    clinicas = [('', '')] + [(clin.clinica, clin.clinica) for clin in CatClinicas.query.filter_by(est_clin='A', interna='S').order_by(CatClinicas.clinica)]
    return jsonify(clinicas)

@catalogos_bp.route('/actualizar_menus', methods=['GET'])
def actualizar_menus():
    menus = [('', '')] + [(men.menu, men.menu) for men in CatMenus.query.filter_by(est_menu='A').order_by(CatMenus.menu)]
    return jsonify(menus)

@catalogos_bp.route('/actualizar_submenus', methods=['GET'])
def actualizar_submenus():
    submenus = [('', '')] + [(submen.submenu, submen.submenu) for submen in CatSubMenus.query.filter_by(est_submenu='A').order_by(CatSubMenus.submenu)]
    return jsonify(submenus)

@catalogos_bp.route('/actualizar_tiporoles', methods=['GET'])
def actualizar_tiporoles():
    tiporoles = [('', '')] + [(tiporol.tp_rol, tiporol.tp_rol) for tiporol in CatRoles.query.filter_by(est_rol='A').order_by(CatRoles.tp_rol)]
    return jsonify(tiporoles)

@catalogos_bp.route('/actualizar_opciones_cies', methods=['GET'])
def actualizar_opciones_cies():
    vers_cie = request.args.get('vers_cie')
    opciones_cies = CatEnfermedades.query.with_entities(CatEnfermedades.est_enf).filter(CatEnfermedades.vers_cie == vers_cie).group_by(CatEnfermedades.est_enf).all()
    # Procesa los resultados de la consulta
    opciones_cies_list = [opcion.est_enf for opcion in opciones_cies]
    return jsonify(opciones_cies_list)

@catalogos_bp.route('/buscar_usuario', methods=['GET'])
def buscar_usuario():
    query = request.args.get('query', '').split()

    if not query:
        return jsonify([])  # Devuelve una lista vacía si no hay términos de búsqueda

    # Comenzamos con una intersección vacía
    interseccion_usuarios = None
    
    for palabra in query:
        search_pattern = f"%{palabra}%"
        usuarios = SyUsuarios.query.filter(
            SyUsuarios.est_usuario == 'A',
            func.concat(SyUsuarios.expediente, ': ', SyUsuarios.nombre, ' ', SyUsuarios.paterno, ' ', SyUsuarios.materno).like(search_pattern)
        ).all()
        
        if interseccion_usuarios is None:
            interseccion_usuarios = set(usuarios)
        else:
            interseccion_usuarios &= set(usuarios)
    
    if not interseccion_usuarios:
        return jsonify([])

    resultados = [
        {
            'id': usuario.id_usuario, 
            'usuario': f"{usuario.expediente}: {usuario.nombre} {usuario.paterno} {usuario.materno}",
        } 
        for usuario in interseccion_usuarios
    ]
    return jsonify(resultados)

@catalogos_bp.route('/buscar_escuela', methods=['GET'])
def buscar_escuela():
    query = request.args.get('query', '').split()

    if not query:
        return jsonify([])  # Devuelve una lista vacía si no hay términos de búsqueda
    
    # Comenzamos con una intersección vacía
    interseccion_escuelas = None
    
    for palabra in query:
        # Construir el formato de búsqueda
        search_pattern = f"%{palabra}%"
        escuelas = CatEscuelas.query.filter(CatEscuelas.est_esc == 'A', func.concat(CatEscuelas.sigls_esc, ': ' , CatEscuelas.escuela).like(search_pattern)).all()

        if interseccion_escuelas is None:
            interseccion_escuelas = set(escuelas)
        else:
            interseccion_escuelas &= set(escuelas)

    if not interseccion_escuelas:
        return jsonify([])            

    resultados = [
        {
            'id': escuela.id_esc, 
            'escuela': f'{escuela.sigls_esc}: {escuela.escuela}',
        } 
        for escuela in interseccion_escuelas
        ]
    return jsonify(resultados)

@catalogos_bp.route('/buscar_especialidad', methods=['GET'])
def buscar_especialidad():
    query = request.args.get('query', '').split()

    if not query:
        return jsonify([])  # Devuelve una lista vacía si no hay términos de búsqueda
    
    # Comenzamos con una intersección vacía
    interseccion_especialidades = None
    
    for palabra in query:
        # Construir el formato de búsqueda
        search_pattern = f"%{palabra}%"
        especialidades = CatEspecialidades.query.filter(CatEspecialidades.est_espec == 'A', (CatEspecialidades.especialidad).like(search_pattern) ).all()

        if interseccion_especialidades is None:
            interseccion_especialidades = set(especialidades)
        else:
            interseccion_especialidades &= set(especialidades)

    if not interseccion_especialidades:
        return jsonify([])            

    resultados = [
        {
            'id': especialidad.id_espec, 
            'especialidad': f"{especialidad.especialidad}",
        } 
        for especialidad in interseccion_especialidades
        ]
    return jsonify(resultados)

@catalogos_bp.route('/buscar_servicio', methods=['GET'])
def buscar_servicio():
    query = request.args.get('query', '').split()

    if not query:
        return jsonify([])  # Devuelve una lista vacía si no hay términos de búsqueda
    
    # Comenzamos con una intersección vacía
    interseccion_servicios = None
    
    for palabra in query:
        # Construir el formato de búsqueda
        search_pattern = f"%{palabra}%"
        servicios = CatServiciosClin.query.filter(CatServiciosClin.est_serviciosclin == 'A', (CatServiciosClin.serviciosclin).like(search_pattern) ).all()

        if interseccion_servicios is None:
            interseccion_servicios = set(servicios)
        else:
            interseccion_servicios &= set(servicios)

    if not interseccion_servicios:
        return jsonify([])            

    resultados = [
        {
            'id': servicio.id_serviciosclin, 
            'servicio': f'{servicio.serviciosclin}',
        } 
        for servicio in interseccion_servicios
        ]
    return jsonify(resultados)

@catalogos_bp.route('/buscar_clinica', methods=['GET'])
def buscar_clinica():
    query = request.args.get('query', '').split()

    if not query:
        return jsonify([])  # Devuelve una lista vacía si no hay términos de búsqueda
    
    # Comenzamos con una intersección vacía
    interseccion_clinicas = None
    
    for palabra in query:
        # Construir el formato de búsqueda
        search_pattern = f"%{palabra}%"
        clinicas = CatClinicas.query.filter(
            CatClinicas.est_clin == 'A',
            CatClinicas.interna == 'S',
            CatClinicas.clinica.like(search_pattern)
        ).all()

        if interseccion_clinicas is None:
            interseccion_clinicas = set(clinicas)
        else:
            interseccion_clinicas &= set(clinicas)

    if not interseccion_clinicas:
        return jsonify([])            

    resultados = [
        {
            'id': clinica.id_clin, 
            'clinica': f'{clinica.clinica}',
        } 
        for clinica in interseccion_clinicas
    ]
    return jsonify(resultados)

@catalogos_bp.route('/buscar_consultorio', methods=['GET'])
def buscar_consultorio():
    query = request.args.get('query', '').split()
    clinica = request.args.get('clinica')

    if not query:
        return jsonify([])  # Devuelve una lista vacía si no hay términos de búsqueda
    
    # Comenzamos con una intersección vacía
    interseccion_consultorios = None
    
    for palabra in query:
        # Construir el formato de búsqueda
        search_pattern = f"%{palabra}%"
        consultorios = CatConsultorios.query.filter(
            CatConsultorios.est_consult == 'A',
            CatConsultorios.clinica == clinica,
            func.concat(CatConsultorios.no_consult, ': ', CatConsultorios.consult, ' (', CatConsultorios.turno, ')').like(search_pattern)
            
        ).all()

        if interseccion_consultorios is None:
            interseccion_consultorios = set(consultorios)
        else:
            interseccion_consultorios &= set(consultorios)

    if not interseccion_consultorios:
        return jsonify([])            

    resultados = [
        {
            'id': consultorio.id_consult, 
            'consultorio': f'{consultorio.no_consult}: {consultorio.consult} ({consultorio.turno})',
        } 
        for consultorio in interseccion_consultorios
    ]
    return jsonify(resultados)

@catalogos_bp.route('/buscar_consultorio2', methods=['GET'])
def buscar_consultorio2():
    query = request.args.get('query', '').split()
    id_clinica = request.args.get('id_clinica')

    if not query:
        return jsonify([])  # Devuelve una lista vacía si no hay términos de búsqueda
    
    # Comenzamos con una intersección vacía
    interseccion_consultorios2 = None
    
    for palabra in query:
        # Construir el formato de búsqueda
        search_pattern = f"%{palabra}%"
        consultorios2 = CatConsultorios.query.filter(
            CatConsultorios.est_consult == 'A',
            CatConsultorios.id_clin == id_clinica,
            func.concat(CatConsultorios.no_consult, ': ', CatConsultorios.consult, ' (', CatConsultorios.turno, ')').like(search_pattern)
            
        ).all()

        if interseccion_consultorios2 is None:
            interseccion_consultorios2 = set(consultorios2)
        else:
            interseccion_consultorios2 &= set(consultorios2)

    if not interseccion_consultorios2:
        return jsonify([])            

    resultados = [
        {
            'id': consultorio2.id_consult, 
            'consultorio2': f'{consultorio2.no_consult}: {consultorio2.consult} ({consultorio2.turno})',
        } 
        for consultorio2 in interseccion_consultorios2
    ]
    return jsonify(resultados)

@catalogos_bp.route('/buscar_rol', methods=['GET'])
def buscar_rol():
    query = request.args.get('query', '').split()

    if not query:
        return jsonify([])  # Devuelve una lista vacía si no hay términos de búsqueda
    
    # Comenzamos con una intersección vacía
    interseccion_roles = None
    
    for palabra in query:
        # Construir el formato de búsqueda
        search_pattern = f"%{palabra}%"
        roles = CatRoles.query.filter(CatRoles.est_rol == 'A', func.concat(CatRoles.tp_rol, ': ' , CatRoles.rol).like(search_pattern)).all()

        if interseccion_roles is None:
            interseccion_roles = set(roles)
        else:
            interseccion_roles &= set(roles)

    if not interseccion_roles:
        return jsonify([])            

    resultados = [
        {
            'id': rol.id_rol, 
            'rol': f'{rol.tp_rol}: {rol.rol}',
        } 
        for rol in interseccion_roles
        ]
    return jsonify(resultados)

@catalogos_bp.route('/buscar_tpautorizador', methods=['GET'])
def buscar_tpautorizador():
    query = request.args.get('query', '').split()

    if not query:
        return jsonify([])  # Devuelve una lista vacía si no hay términos de búsqueda
    
    # Comenzamos con una intersección vacía
    interseccion_tpautorizadores = None
    
    for palabra in query:
        # Construir el formato de búsqueda
        search_pattern = f"%{palabra}%"
        tpautorizadores = CatTpAutorizador.query.filter(CatTpAutorizador.est_tpautorizador == 'A', func.concat(CatTpAutorizador.id_tpautorizador, ': ' , CatTpAutorizador.tpautorizador).like(search_pattern)).all()

        if interseccion_tpautorizadores is None:
            interseccion_tpautorizadores = set(tpautorizadores)
        else:
            interseccion_tpautorizadores &= set(tpautorizadores)

    if not interseccion_tpautorizadores:
        return jsonify([])            

    resultados = [
        {
            'id': tpautorizador.id_tpautorizador, 
            'tpautorizador': f'{tpautorizador.id_tpautorizador}: {tpautorizador.tpautorizador}',
        } 
        for tpautorizador in interseccion_tpautorizadores
        ]
    return jsonify(resultados)

@catalogos_bp.route('/api/datos_cp')
def obtener_datos():
    if datos_estructurados is None:
        return jsonify({'error': 'Datos aún no cargados'}), 503
    return jsonify(datos_estructurados)

# Almacenar el DataFrame en memoria para la paginación
df_global = None
total_errores = None
version = None

@catalogos_bp.route('/upload_file', methods=['POST'])
def upload_file():
    global df_global
    global version
    global total_errores
    try:
        archivo_excel = request.files['file']
        version = request.form.get('version')
        df_global = pd.read_excel(io.BytesIO(archivo_excel.read()), usecols=[0, 1])

        # Limpieza de espacios en blanco
        df_global[df_global.columns[0]] = df_global[df_global.columns[0]].apply(lambda x: re.sub(r'\s+', ' ', str(x).strip()) if not pd.isna(x) else "")
        df_global[df_global.columns[1]] = df_global[df_global.columns[1]].apply(lambda x: re.sub(r'\s+', ' ', str(x).strip()) if not pd.isna(x) else "")

        # Inicializar la columna de errores
        df_global['version'] = version
        df_global['error'] = ""

        # Verificar claves vacías
        if (df_global[df_global.columns[0]] == "").any():
            df_global.loc[df_global[df_global.columns[0]] == "", 'error'] = "Clave vacía. "

        if (df_global[df_global.columns[0]].str.len() > 8).any():
            df_global.loc[df_global[df_global.columns[0]].str.len() > 8, 'error'] += "Clave demasiado larga. Máximo 8 caracteres. "

        df_clave_empty = df_global[df_global[df_global.columns[0]] != ""]

        if df_clave_empty.duplicated(subset=[df_clave_empty.columns[0]], keep=False).any():
            df_global.loc[df_global[df_global.columns[0]].isin(df_clave_empty[df_clave_empty.duplicated(subset=[df_clave_empty.columns[0]], keep=False)][df_clave_empty.columns[0]]), 'error'] += "Clave duplicada. "

        # Verificar claves duplicadas
        #if df_global.duplicated(subset=[df_global.columns[0]], keep=False).any():
        #    df_global.loc[df_global.duplicated(subset=[df_global.columns[0]], keep=False), 'error'] += "Clave duplicada. "

        # Verificar descripciones vacías
        if (df_global[df_global.columns[1]] == "").any():
            df_global.loc[df_global[df_global.columns[1]] == "", 'error'] += "Descripción de enfermedad vacía. "

        if (df_global[df_global.columns[1]].str.len() > 400).any():
            df_global.loc[df_global[df_global.columns[1]].str.len() > 400, 'error'] += "Descripción de enfermedad demasiado larga. Máximo 400 caracteres. "

        df_desc_empty = df_global[df_global[df_global.columns[1]] != ""]

        if df_desc_empty.duplicated(subset=[df_desc_empty.columns[1]], keep=False).any():
            df_global.loc[df_global[df_global.columns[1]].isin(df_desc_empty[df_desc_empty.duplicated(subset=[df_desc_empty.columns[1]], keep=False)][df_desc_empty.columns[1]]), 'error'] += "Descripción de enfermedad duplicada. "

        # Verificar descripciones duplicadas
        #if df_global.duplicated(subset=[df_global.columns[1]], keep=False).any():
        #    df_global.loc[df_global.duplicated(subset=[df_global.columns[1]], keep=False), 'error'] += "Descripción de enfermedad duplicada. "

        df_errores = df_global[df_global['error'] != ""]
        total_errores = len(df_errores)

        # Cambiar los nombres de las columnas
        df_global.columns = ['CLAVE', 'DESCRIPCIÓN DE ENFERMEDAD', 'VERSIÓN CIE', 'TIPO DE ERROR']

        return get_paginated_data(1)
    except Exception as e:
        return jsonify({"error": str(e)})

@catalogos_bp.route('/get_data', methods=['GET'])
def get_data():
    try:
        page = int(request.args.get('page', 1))
        return get_paginated_data(page)
    except Exception as e:
        return jsonify({"error": str(e)})
    
def get_paginated_data(page, per_page=5):
    global df_global
    global version
    global total_errores
    if df_global is None:
        return jsonify({"error": "No data available"})
    if total_errores is None:
        return jsonify({"error": "Total de errores no disponible"})

    start = (page - 1) * per_page
    end = start + per_page
    df_page = df_global.iloc[start:end]
    total_records = len(df_global)
    total_pages = math.ceil(total_records / per_page)

    data = {
        "total_records": total_records,
        "total_errores": total_errores,
        "total_pages": total_pages,
        "current_page": page,
        "version": version,
        "rows": []
    }

    for index, row in df_page.iterrows():
        col0 = row[0]
        col1 = row[1]
        error = row[3]

        data["rows"].append({
            "index": index,
            "col0": col0,
            "col1": col1,
            "error": error
        })

    return jsonify(data)

@catalogos_bp.route('/download_errors', methods=['GET'])
def download_errors():
    global df_global

    if df_global is None or df_global.empty:
        return jsonify({"error": "No errors available to download"})
    
    # Crear un buffer de bytes en memoria para el archivo Excel
    output = io.BytesIO()

    # Crear un objeto ExcelWriter usando el buffer de bytes
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_global.to_excel(writer, sheet_name='Errores', index=False)

    # Mover el cursor al inicio del buffer
    output.seek(0)

    # Cargar el archivo Excel desde el buffer
    workbook = load_workbook(output)
    worksheet = workbook.active

    # Ajustar el ancho de las columnas
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width    

    # Definir el estilo de relleno rojo
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # Aplicar el relleno rojo a las filas donde la columna 'ERROR' no está vacía
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        error_cell = row[3]
        if error_cell.value:
            for cell in row:
                cell.fill = red_fill

    # Guardar el archivo con los estilos aplicados en el buffer
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Enviar el archivo como respuesta
    return send_file(
        output,
        download_name='errores.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

########################---------------------------------CARGAR DATOS PARA FARMACIA--------------------------------------############################
'''catalogos_bp.route('/upload_file_farmacia', methods=['POST'])
def upload_file():
    global df_global
    global total_errores
    try:
        archivo_excel = request.files['file']
        df_global = pd.read_excel(io.BytesIO(archivo_excel.read()), usecols=[0, 1])

        # Limpieza de espacios en blanco
        df_global[df_global.columns[0]] = df_global[df_global.columns[0]].apply(lambda x: re.sub(r'\s+', ' ', str(x).strip()) if not pd.isna(x) else "")
        df_global[df_global.columns[1]] = df_global[df_global.columns[1]].apply(lambda x: re.sub(r'\s+', ' ', str(x).strip()) if not pd.isna(x) else "")

        # Inicializar la columna de errores
        df_global['error'] = ""

        # Verificar claves vacías
        if (df_global[df_global.columns[0]] == "").any():
            df_global.loc[df_global[df_global.columns[0]] == "", 'error'] = "Clave vacía. "

        if (df_global[df_global.columns[0]].str.len() > 8).any():
            df_global.loc[df_global[df_global.columns[0]].str.len() > 8, 'error'] += "Clave demasiado larga. Máximo 8 caracteres. "

        df_clave_empty = df_global[df_global[df_global.columns[0]] != ""]

        if df_clave_empty.duplicated(subset=[df_clave_empty.columns[0]], keep=False).any():
            df_global.loc[df_global[df_global.columns[0]].isin(df_clave_empty[df_clave_empty.duplicated(subset=[df_clave_empty.columns[0]], keep=False)][df_clave_empty.columns[0]]), 'error'] += "Clave duplicada. "

        # Verificar claves duplicadas
        #if df_global.duplicated(subset=[df_global.columns[0]], keep=False).any():
        #    df_global.loc[df_global.duplicated(subset=[df_global.columns[0]], keep=False), 'error'] += "Clave duplicada. "

        # Verificar descripciones vacías
        if (df_global[df_global.columns[1]] == "").any():
            df_global.loc[df_global[df_global.columns[1]] == "", 'error'] += "Descripción de enfermedad vacía. "

        if (df_global[df_global.columns[1]].str.len() > 400).any():
            df_global.loc[df_global[df_global.columns[1]].str.len() > 400, 'error'] += "Descripción de enfermedad demasiado larga. Máximo 400 caracteres. "

        df_desc_empty = df_global[df_global[df_global.columns[1]] != ""]

        if df_desc_empty.duplicated(subset=[df_desc_empty.columns[1]], keep=False).any():
            df_global.loc[df_global[df_global.columns[1]].isin(df_desc_empty[df_desc_empty.duplicated(subset=[df_desc_empty.columns[1]], keep=False)][df_desc_empty.columns[1]]), 'error'] += "Descripción de enfermedad duplicada. "

        # Verificar descripciones duplicadas
        #if df_global.duplicated(subset=[df_global.columns[1]], keep=False).any():
        #    df_global.loc[df_global.duplicated(subset=[df_global.columns[1]], keep=False), 'error'] += "Descripción de enfermedad duplicada. "

        df_errores = df_global[df_global['error'] != ""]
        total_errores = len(df_errores)

        # Cambiar los nombres de las columnas
        df_global.columns = ['CLAVE', 'MEDICAMENTO', 'DESCRIPCIÓN', 'UNIDADES', 'SURTIDO']

        return get_paginated_data(1)
    except Exception as e:
        return jsonify({"error": str(e)})'''
#######################-----------------------------------FIN CARGA DATOS PARA FARMACIA--------------------------------------############################


########################---------------------------------REPORTES--------------------------------------------------------############################
@catalogos_bp.route('/repnuevo', methods=['POST'])
def reporte_nuevo():
    tabla = request.form.get('tabla')
    filas = tabla.splitlines()
    valor = request.form.get('cat', '')
    array_aux = [re.split(r'\s{2,}', line) for line in filas]
    array = []

    for row in array_aux:
        array.append(row[0].split('\t'))

    estiloHoja = getSampleStyleSheet()
    thead = estiloHoja["Normal"]
    thead.alignment = TA_CENTER
    tbody = estiloHoja["BodyText"]
    tbody.alignment = TA_LEFT
    
    ruta_archivo = os.path.join(os.getcwd(), "reports", "Reporte " + request.form.get('nom_rep').lower() + ".pdf")
    #ruta_archivo = "./reports/Reporte " + request.form.get('nom_rep').lower() + ".pdf"

    if valor in ['hospitales', 'laboratorios', 'estudiosmed', 'clinicas', 'consultorios', 'areas', 'medicosclin', 'medicamentos']:
        PAGE_WIDTH, PAGE_HEIGHT = landscape(letter)
        doc = SimpleDocTemplate(ruta_archivo, pagesize=landscape(letter), leftMargin=20, rightMargin=20, topMargin=90, bottomMargin=50)
    else:
        PAGE_WIDTH, PAGE_HEIGHT = portrait(letter)
        doc = SimpleDocTemplate(ruta_archivo, pagesize=portrait(letter), leftMargin=20, rightMargin=20, topMargin=90, bottomMargin=50)
    
    datos = []
    for fila in array:
        fila_para_tabla = [Paragraph(str(celda), tbody) for celda in fila]
        datos.append(fila_para_tabla)


    if valor == 'areas': colWidths=[5.0 * inch, 3.0 * inch, 1.1 * inch]
    elif valor == 'calidadlab': colWidths=[2.0 * inch, 2.5 * inch, 2.0 * inch]
    elif valor == 'cie10': colWidths=[0.8 * inch , 2.5 * inch, 1.2 * inch, 1.5 * inch, 1.1 * inch]
    elif valor == 'cie11': colWidths=[0.8 * inch , 2.5 * inch, 1.2 * inch, 1.5 * inch, 1.1 * inch]
    elif valor == 'clinicas': colWidths=[1.0 * inch, 0.7 * inch, 1.0 * inch, 2.3 * inch, 2.5 * inch, 2.3 * inch, 1.1 * inch ]
    elif valor == 'consultorios': colWidths=[0.8 * inch , 1.5 * inch, 2.0 * inch, 2.5 * inch, 1.1 * inch]
    elif valor == 'edocivil': colWidths=[4.0 * inch, 4.0 * inch]
    elif valor == 'enfermedades': colWidths=[1.3 * inch, 3.0 * inch, 1.5 * inch, 1.1 * inch]
    elif valor == 'escolaridad': colWidths=[ 3.5 * inch, 3.5 * inch]
    elif valor == 'escuelas': colWidths=[ 3.5 * inch, 2.0 * inch, 1.4 * inch] 
    elif valor == 'especialidades': colWidths=[4.5 * inch, 2.5 * inch]
    elif valor == 'estudiosmed': colWidths=[2.0 * inch, 0.8 * inch, 1.4 * inch, 1.5 * inch, 1.1 * inch, 1.1 * inch, 0.8 * inch, 0.8 * inch, 1.1 * inch]
    elif valor == 'gpomedic': colWidths=[3.5 * inch, 3.5 * inch]
    elif valor == 'hospitales': colWidths=[2.0 * inch, 5.0 * inch, 1.5 * inch, 1.1 * inch]
    elif valor == 'laboratorios': colWidths=[1.5 * inch, 1.5 * inch, 3.5 * inch, 1.5 * inch, 1.1 * inch]
    elif valor == 'ocupaciones': colWidths=[4.0 * inch, 3.5 * inch]
    elif valor == 'origencons': colWidths=[2.0 * inch, 3.5 * inch, 1.1 * inch]
    elif valor == 'parentescos': colWidths=[1.5 * inch, 3.5 * inch, 1.5 * inch]
    elif valor == 'pases': colWidths=[3.0 * inch, 3.0 * inch]
    elif valor == 'serviciosclin': colWidths=[4.5 * inch, 1.5 * inch]
    elif valor == 'tipo_consulta': colWidths=[4.0 * inch, 3.0 * inch]
    elif valor == 'tipo_hospn': colWidths=[4.5 * inch, 2.0 * inch]
    elif valor == 'tpareas': colWidths=[3.0 * inch, 3.0 * inch]    
    elif valor == 'tpautorizacion': colWidths=[ 2.0 * inch, 4.0 * inch, 1.1 * inch]
    elif valor == 'tpautorizador': colWidths=[2.0 * inch, 2.5 * inch, 2.0 * inch]
    elif valor == 'tpbajas': colWidths=[1.0 * inch, 4.8 * inch, 1.1 * inch]
    elif valor == 'tpcitas': colWidths=[3.0 * inch, 3.0 * inch]
    elif valor == 'tplicencia': colWidths=[4.5 * inch, 3.0 * inch]
    elif valor == 'tpsanguineo': colWidths=[4.0 * inch, 2.0 * inch]
    elif valor == 'turnos': colWidths=[4.5 * inch, 2.0 * inch] 
    elif valor == 'medicamentos': colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch]     


    def dibujar_encabezado(canvas, doc):
        canvas.saveState()
        pdfmetrics.registerFont(TTFont('stc', './app/static/css/METRO-DF.TTF'))
        titulo = request.form.get('nom_rep')
        image_left_path = './app/static/img/Logo_Dependencia.jpg'
        image_right_path = './app/static/img/Logo_CDMX.jpg'
        y2 = PAGE_HEIGHT - 1.5 * inch

        canvas.drawImage(image_left_path, 2.18 * inch, PAGE_HEIGHT - 1.4 * inch, width = 3.2 * inch, preserveAspectRatio=True)
        canvas.drawImage(image_right_path, 0.2 * inch, PAGE_HEIGHT - 2.18 * inch, width = 1.9 * inch, preserveAspectRatio=True)
        canvas.setFont("stc", 9)
        canvas.drawCentredString(PAGE_WIDTH - 1.97 * inch, y2 + 1.14 * inch, "GERENCIA DE SALUD Y BIENESTAR SOCIAL")
        canvas.drawCentredString(PAGE_WIDTH - 1.94 * inch, y2 + 0.98 * inch, "SISTEMA DE INFORMACIÓN DE REGISTRO")
        canvas.drawCentredString(PAGE_WIDTH - 1.46 * inch, y2 + 0.85 * inch, "ELECTRÓNICO PARA LA SALUD")
        canvas.setLineWidth(1)
        canvas.line(PAGE_WIDTH - 3.75 * inch , y2 + 1.1 * inch, PAGE_WIDTH - 0.224 * inch, y2 + 1.1 * inch)
        canvas.setFont("stc", 12)
        canvas.drawCentredString(PAGE_WIDTH / 2.0, PAGE_HEIGHT - 70, titulo)

        # Crear el estilo de texto para los encabezados de la tabla
        estiloHoja = getSampleStyleSheet()
        thead = estiloHoja["Normal"]
        thead.fontName = 'stc'  # Usar la fuente registrada 'stc'
        thead.alignment = TA_LEFT

        # Dibujar el encabezado de la tabla con estilo
        encabezado_tabla = [[Paragraph(col, thead) for col in array[0]]]  # Crear los encabezados como Paragraphs

        table = Table(encabezado_tabla, colWidths=colWidths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), '#d0cece'),  # Fondo gris claro para los encabezados
            ('TEXTCOLOR', (0, 0), (-1, 0), 'black'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'stc'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
        ]))
        
        # Ajustar la posición de la tabla en el canvas
        w, h = table.wrapOn(canvas, PAGE_WIDTH - 51.8, 100)
        canvas.setFont("stc", 2)
        table.drawOn(canvas, PAGE_WIDTH/2 - (w/2), PAGE_HEIGHT - 78 - h)  # Ajusta la posición según sea necesario
            
        canvas.restoreState()

    def dibujar_pie(canvas, doc):
        canvas.saveState()
        fecha_impresion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        x = PAGE_WIDTH - 1.5 * inch
        y = 0.4 * inch
        canvas.setFont("stc", 8)
        canvas.drawCentredString(x, y, f"Página {doc.page} - {fecha_impresion}")
        canvas.restoreState()

    def myFirstPage(canvas, doc):
        dibujar_encabezado(canvas, doc)
        dibujar_pie(canvas, doc)

    def myLaterPages(canvas, doc):
        dibujar_encabezado(canvas, doc)
        dibujar_pie(canvas, doc)

    tabla_estilo = TableStyle([
        # Estilos para el cuerpo de la tabla
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), 
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Fuente normal para los datos
        ('FONTSIZE', (0, 1), (-1, -1), 2),  # Tamaño de fuente 6 para los datos
        ('LINEBELOW', (0, 0), (-1, -1), 0.3, colors.black, None, (1, 1, 0)),  # Líneas debajo de cada fila de datos
    ])  

    story = []
    story.append(Table(datos[1:], colWidths=colWidths, style=tabla_estilo))

    doc.build(story, onFirstPage=myFirstPage, onLaterPages=myLaterPages)

    #return send_file(ruta_archivo, as_attachment=True)
    return send_file(ruta_archivo, as_attachment=True, mimetype='application/pdf')
