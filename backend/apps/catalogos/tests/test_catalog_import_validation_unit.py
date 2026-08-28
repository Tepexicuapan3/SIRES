"""
Unit tests for the generic catalog-import engine (`import-catalogos-excel`):
- `parse_and_validate` per-rule validation (ID / Nombre / Clave / Activo)
- `build_template` header generation per catalog spec

These hit the DB only where the validated rule itself requires a DB lookup
(the ID-collision check via `spec.model.objects.filter(pk__in=...)`); every
other case is pure in-memory .xlsx parsing, matching the design's "Testing
Strategy" table (openpyxl-generated .xlsx via BytesIO, no DB for pure rule
checks).
"""

import io

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase

from apps.catalogos.imports.registry import CATALOG_IMPORT_REGISTRY
from apps.catalogos.models import Especialidades
from apps.catalogos.services.catalog_import_service import (
    ImportFileError,
    build_template,
    parse_and_validate,
)

DISABILITIES_SPEC = CATALOG_IMPORT_REGISTRY["disabilities"]
SPECIALTIES_SPEC = CATALOG_IMPORT_REGISTRY["specialties"]
SCHOOLS_SPEC = CATALOG_IMPORT_REGISTRY["schools"]

DISABILITIES_HEADERS = ["ID", "Clave", "Nombre", "Activo"]
SPECIALTIES_HEADERS = ["ID", "Nombre", "Activo"]
SCHOOLS_HEADERS = ["ID", "Clave", "Nombre", "Activo"]


def _xlsx_upload(headers, rows, filename="import.xlsx"):
    """Builds an in-memory .xlsx (openpyxl -> BytesIO) and wraps it as the
    kind of UploadedFile `request.FILES.get("file")` hands the service."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    buffer = io.BytesIO()
    wb.save(buffer)
    return SimpleUploadedFile(
        filename,
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class BuildTemplateTests(TestCase):
    def test_disabilities_template_headers_and_sample_rows(self):
        content = build_template(DISABILITIES_SPEC)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, DISABILITIES_HEADERS)
        self.assertEqual(ws.cell(row=2, column=1).value, 1)

    def test_schools_template_headers(self):
        content = build_template(SCHOOLS_SPEC)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        headers = [cell.value for cell in wb.active[1]]
        self.assertEqual(headers, SCHOOLS_HEADERS)

    def test_specialties_template_has_no_clave_column(self):
        content = build_template(SPECIALTIES_SPEC)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        headers = [cell.value for cell in wb.active[1]]

        self.assertEqual(headers, SPECIALTIES_HEADERS)
        self.assertNotIn("Clave", headers)
        self.assertNotIn("Codigo", headers)
        self.assertNotIn("Código", headers)


class ParseAndValidateFileLevelTests(TestCase):
    def test_wrong_extension_raises_import_file_error(self):
        file = SimpleUploadedFile("import.txt", b"not an excel file", content_type="text/plain")

        with self.assertRaises(ImportFileError) as ctx:
            parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(ctx.exception.code, "IMPORT_FILE_INVALID")

    def test_headers_mismatch_raises(self):
        file = _xlsx_upload(["ID", "OtraColumna", "Activo"], [["1", "x", "Si"]])

        with self.assertRaises(ImportFileError) as ctx:
            parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(ctx.exception.code, "IMPORT_HEADERS_MISMATCH")

    def test_specialties_rejects_a_clave_column_it_does_not_own(self):
        # Especialidades NO lleva Clave: un archivo con esa columna de mas
        # (aunque el resto coincida) debe rechazarse como headers-mismatch.
        file = _xlsx_upload(
            ["ID", "Clave", "Nombre", "Activo"],
            [["1", "X01", "Cardiología", "Si"]],
        )

        with self.assertRaises(ImportFileError) as ctx:
            parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(ctx.exception.code, "IMPORT_HEADERS_MISMATCH")

    def test_too_many_rows_raises(self):
        rows = [[str(i), f"Especialidad {i}", "Si"] for i in range(1, SPECIALTIES_SPEC.max_rows + 2)]
        file = _xlsx_upload(SPECIALTIES_HEADERS, rows)

        with self.assertRaises(ImportFileError) as ctx:
            parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(ctx.exception.code, "IMPORT_TOO_MANY_ROWS")


class ParseAndValidateIdTests(TestCase):
    def test_blank_id_is_required_error(self):
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["", "Cardiología", "Si"]])
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(result["total_errores"], 1)
        self.assertIn("ID vacío", result["rows"][0]["ERROR"])

    def test_non_integer_id_is_error(self):
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["abc", "Cardiología", "Si"]])
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(result["total_errores"], 1)
        self.assertIn("ID debe ser un número entero", result["rows"][0]["ERROR"])

    def test_decimal_id_is_error(self):
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["1.5", "Cardiología", "Si"]])
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(result["total_errores"], 1)
        self.assertIn("ID debe ser un número entero", result["rows"][0]["ERROR"])

    def test_negative_id_is_error(self):
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["-1", "Cardiología", "Si"]])
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(result["total_errores"], 1)
        self.assertIn("ID no puede ser negativo", result["rows"][0]["ERROR"])

    def test_zero_id_is_valid(self):
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["0", "Cardiología", "Si"]])
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(result["total_errores"], 0)
        self.assertEqual(result["rows"][0]["ERROR"], "")

    def test_duplicate_id_within_file_flags_both_rows(self):
        file = _xlsx_upload(
            SPECIALTIES_HEADERS,
            [
                ["7", "Cardiología", "Si"],
                ["7", "Pediatría", "Si"],
            ],
        )
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(result["total_errores"], 2)
        for row in result["rows"]:
            self.assertIn("ID duplicado en el archivo", row["ERROR"])

    def test_id_colliding_with_existing_db_row_is_error(self):
        Especialidades.objects.create(id=3, name="Ya existe", is_active=True)
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["3", "Cardiología", "Si"]])

        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(result["total_errores"], 1)
        self.assertIn("El ID ya existe en el catálogo", result["rows"][0]["ERROR"])

    def test_valid_id_has_no_error(self):
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["5", "Cardiología", "Si"]])
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(result["total_errores"], 0)
        self.assertEqual(result["rows"][0]["ERROR"], "")

    def test_gaps_are_preserved_never_renumbered(self):
        given_ids = [1, 2, 3, 7, 8, 15, 16, 20]
        rows = [[str(i), f"Especialidad {i}", "Si"] for i in given_ids]
        file = _xlsx_upload(SPECIALTIES_HEADERS, rows)

        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(result["total_errores"], 0)
        self.assertEqual([record["id"] for record in result["records"]], given_ids)


class ParseAndValidateNombreTests(TestCase):
    def test_blank_nombre_is_error(self):
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["1", "", "Si"]])
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertIn("Nombre vacío", result["rows"][0]["ERROR"])

    def test_nombre_over_max_length_specialties_is_error(self):
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["1", "x" * 101, "Si"]])
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertIn("Nombre demasiado largo. Máximo 100 caracteres", result["rows"][0]["ERROR"])

    def test_nombre_at_max_length_specialties_is_valid(self):
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["1", "x" * 100, "Si"]])
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(result["total_errores"], 0)

    def test_nombre_over_max_length_disabilities_is_error(self):
        file = _xlsx_upload(DISABILITIES_HEADERS, [["1", "D01", "x" * 301, "Si"]])
        result = parse_and_validate(DISABILITIES_SPEC, file)

        self.assertIn("Nombre demasiado largo. Máximo 300 caracteres", result["rows"][0]["ERROR"])

    def test_nombre_over_max_length_schools_is_error(self):
        file = _xlsx_upload(SCHOOLS_HEADERS, [["1", "E01", "x" * 101, "Si"]])
        result = parse_and_validate(SCHOOLS_SPEC, file)

        self.assertIn("Nombre demasiado largo. Máximo 100 caracteres", result["rows"][0]["ERROR"])


class ParseAndValidateClaveTests(TestCase):
    def test_clave_is_optional_absent_becomes_empty_string(self):
        file = _xlsx_upload(DISABILITIES_HEADERS, [["1", "", "Discapacidad", "Si"]])
        result = parse_and_validate(DISABILITIES_SPEC, file)

        self.assertEqual(result["total_errores"], 0)
        self.assertEqual(result["records"][0]["code"], "")

    def test_clave_over_max_length_disabilities_is_error(self):
        file = _xlsx_upload(DISABILITIES_HEADERS, [["1", "x" * 11, "Discapacidad", "Si"]])
        result = parse_and_validate(DISABILITIES_SPEC, file)

        self.assertIn("Clave demasiado largo. Máximo 10 caracteres", result["rows"][0]["ERROR"])

    def test_clave_at_max_length_disabilities_is_valid(self):
        file = _xlsx_upload(DISABILITIES_HEADERS, [["1", "x" * 10, "Discapacidad", "Si"]])
        result = parse_and_validate(DISABILITIES_SPEC, file)

        self.assertEqual(result["total_errores"], 0)

    def test_clave_over_max_length_schools_is_error(self):
        file = _xlsx_upload(SCHOOLS_HEADERS, [["1", "x" * 46, "Escuela", "Si"]])
        result = parse_and_validate(SCHOOLS_SPEC, file)

        self.assertIn("Clave demasiado largo. Máximo 45 caracteres", result["rows"][0]["ERROR"])

    def test_specialties_records_never_carry_a_code_field(self):
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["1", "Cardiología", "Si"]])
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertNotIn("code", result["records"][0])


class ParseAndValidateActivoTests(TestCase):
    def test_si_variants_case_and_accent_insensitive(self):
        for value in ["Si", "si", "SI", "Sí", "sí", "SÍ"]:
            file = _xlsx_upload(SPECIALTIES_HEADERS, [["1", "Cardiología", value]])
            result = parse_and_validate(SPECIALTIES_SPEC, file)

            self.assertEqual(result["total_errores"], 0, msg=f"valor={value!r}")
            self.assertIs(result["records"][0]["is_active"], True, msg=f"valor={value!r}")

    def test_no_variants_case_insensitive(self):
        for value in ["No", "no", "NO"]:
            file = _xlsx_upload(SPECIALTIES_HEADERS, [["1", "Cardiología", value]])
            result = parse_and_validate(SPECIALTIES_SPEC, file)

            self.assertEqual(result["total_errores"], 0, msg=f"valor={value!r}")
            self.assertIs(result["records"][0]["is_active"], False, msg=f"valor={value!r}")

    def test_blank_activo_defaults_to_si(self):
        file = _xlsx_upload(SPECIALTIES_HEADERS, [["1", "Cardiología", ""]])
        result = parse_and_validate(SPECIALTIES_SPEC, file)

        self.assertEqual(result["total_errores"], 0)
        self.assertIs(result["records"][0]["is_active"], True)

    def test_invalid_literals_are_rejected_not_treated_as_truthy(self):
        for value in ["true", "1", "yes", "verdadero", "SI "]:
            file = _xlsx_upload(SPECIALTIES_HEADERS, [["1", "Cardiología", value]])
            result = parse_and_validate(SPECIALTIES_SPEC, file)

            if value.strip().lower() == "si":
                # "SI " colapsa a "si" tras el strip/collapse de espacios -> valido
                self.assertEqual(result["total_errores"], 0, msg=f"valor={value!r}")
                continue

            self.assertEqual(result["total_errores"], 1, msg=f"valor={value!r}")
            self.assertIn("Activo debe ser 'Si' o 'No'", result["rows"][0]["ERROR"])
