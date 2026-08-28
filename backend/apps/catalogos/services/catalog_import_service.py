import io
import re
import unicodedata
from typing import Optional

import openpyxl
import pandas as pd

from apps.catalogos.imports.registry import CatalogImportSpec, ImportColumn


class ImportFileError(Exception):
    """Error de nivel-archivo (headers, extensión, límite de filas): 400, ninguna fila se evalúa."""

    def __init__(self, code: str, message: str, details: Optional[dict] = None):
        self.code = code
        self.message = message
        self.details = details or {}
        super().__init__(message)


def build_template(spec: CatalogImportSpec) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = spec.slug

    for col_idx, column in enumerate(spec.columns, start=1):
        ws.cell(row=1, column=col_idx, value=column.header)

    for row_idx, sample_row in enumerate(spec.sample_rows, start=2):
        for col_idx, value in enumerate(sample_row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _clean_text(value) -> str:
    text = "" if value is None else str(value)
    if text.strip().lower() == "nan":
        text = ""
    return re.sub(r"\s+", " ", text.strip())


def _strip_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(c for c in normalized if not unicodedata.combining(c))


def _parse_int_id(raw) -> tuple:
    text = _clean_text(raw)
    if text == "":
        return None, "ID vacío. "
    try:
        as_float = float(text)
    except ValueError:
        return None, "ID debe ser un número entero. "
    if not as_float.is_integer():
        return None, "ID debe ser un número entero. "
    value = int(as_float)
    if value < 0:
        return None, "ID no puede ser negativo. "
    return value, ""


def _parse_text(raw, column: ImportColumn) -> tuple:
    text = _clean_text(raw)
    error = ""
    if column.required and text == "":
        error += f"{column.header} vacío. "
    if column.max_length is not None and len(text) > column.max_length:
        error += f"{column.header} demasiado largo. Máximo {column.max_length} caracteres. "
    return text, error


def _parse_bool_si_no(raw) -> tuple:
    # "Si"/"Sí"/"No" case- y acento-insensible; vacío => Si. Cualquier otro literal
    # ("true", "1", "SI ") es error explícito, nunca se interpreta como truthy.
    text = _strip_accents(_clean_text(raw)).lower()
    if text == "" or text == "si":
        return True, ""
    if text == "no":
        return False, ""
    return True, "Activo debe ser 'Si' o 'No'. "


def _read_dataframe(spec: CatalogImportSpec, file) -> pd.DataFrame:
    filename = (getattr(file, "name", "") or "").lower()
    if not filename.endswith((".xlsx", ".xls")):
        raise ImportFileError(
            code="IMPORT_FILE_INVALID",
            message="El archivo debe ser .xlsx o .xls.",
        )

    try:
        df = pd.read_excel(file, dtype=str, engine="openpyxl" if filename.endswith(".xlsx") else None)
    except Exception as exc:
        raise ImportFileError(
            code="IMPORT_FILE_INVALID",
            message="No se pudo leer el archivo. Verifique que sea un Excel válido.",
        ) from exc

    df = df.fillna("")

    expected_headers = [column.header for column in spec.columns]
    actual_headers = [str(c).strip() for c in df.columns.tolist()]
    if actual_headers != expected_headers:
        raise ImportFileError(
            code="IMPORT_HEADERS_MISMATCH",
            message="Las columnas del archivo no coinciden con la plantilla del catálogo.",
            details={"expected": expected_headers, "actual": actual_headers},
        )

    if len(df) > spec.max_rows:
        raise ImportFileError(
            code="IMPORT_TOO_MANY_ROWS",
            message=f"El archivo excede el máximo de {spec.max_rows} filas.",
        )

    return df


def parse_and_validate(spec: CatalogImportSpec, file) -> dict:
    """
    Única función de validación, reusada tal cual por preview y por confirm — el
    confirm NUNCA confía en un campo ERROR mandado por el cliente, siempre re-corre
    esto contra el archivo re-subido.
    """
    df = _read_dataframe(spec, file)
    total_records = len(df)

    id_column = next(column for column in spec.columns if column.kind == "int_id")
    other_columns = [column for column in spec.columns if column.kind != "int_id"]

    parsed_ids = []
    id_errors = []
    for raw in df[id_column.header]:
        value, error = _parse_int_id(raw)
        parsed_ids.append(value)
        id_errors.append(error)

    id_counts: dict = {}
    for value in parsed_ids:
        if value is not None:
            id_counts[value] = id_counts.get(value, 0) + 1

    valid_ids = [value for value in parsed_ids if value is not None]
    existing_ids = set(
        spec.model.objects.filter(pk__in=valid_ids).values_list("pk", flat=True)
    ) if valid_ids else set()

    display_rows = []
    records = []
    total_errores = 0

    for row_idx in range(total_records):
        row_error = id_errors[row_idx]
        parsed_id = parsed_ids[row_idx]

        if parsed_id is not None and id_counts[parsed_id] > 1:
            row_error += "ID duplicado en el archivo. "
        if parsed_id is not None and parsed_id in existing_ids:
            row_error += "El ID ya existe en el catálogo. "

        display_row = {
            id_column.header: (
                parsed_id if parsed_id is not None
                else _clean_text(df.iloc[row_idx][id_column.header])
            )
        }
        record = {"id": parsed_id}

        for column in other_columns:
            raw_value = df.iloc[row_idx][column.header]
            if column.kind == "text":
                value, col_error = _parse_text(raw_value, column)
                display_row[column.header] = value
                record[column.field] = value
            elif column.kind == "bool_si_no":
                value, col_error = _parse_bool_si_no(raw_value)
                display_row[column.header] = "Si" if value else "No"
                record[column.field] = value
            else:  # pragma: no cover - guarda contra specs mal configurados
                raise ValueError(f"Tipo de columna no soportado: {column.kind}")
            row_error += col_error

        display_row["ERROR"] = row_error
        display_rows.append(display_row)
        records.append(record)

        if row_error:
            total_errores += 1

    return {
        "total_records": total_records,
        "total_errores": total_errores,
        "rows": display_rows,
        "records": records,
    }
