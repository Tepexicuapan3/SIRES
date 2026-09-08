"""
Servicio de import masivo de usuarios por Excel.

Sigue el mismo estilo que `apps/catalogos/services/catalog_import_service.py`
(openpyxl para escribir la plantilla, pandas para leer/validar), pero NO
reusa el repositorio generico de catalogos (`CatalogImportRepository.bulk_insert`)
porque dar de alta un usuario no es un insert simple: toca `SyUsuario` +
`DetUsuario` + `RelUsuarioRol`, genera password temporal y potencialmente
envia correo de credenciales -- eso lo resuelve
`apps.administracion.use_cases.users.create_user.CreateUserUseCase`.

Este modulo solo se ocupa de leer el .xlsx y devolver, por fila, los datos
parseados + la lista de errores. NUNCA persiste nada (ni preview ni confirm
llaman a este modulo para escribir -- ver `import_users.py`).
"""

import io
import re
import unicodedata
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from apps.authentication.models import SyUsuario
from apps.catalogos.models import CatTipoPersonal, Roles

HEADERS = [
    "Usuario",
    "Nombre(s)",
    "Apellido Paterno",
    "Apellido Materno",
    "Correo",
    "No. Expediente SERMED",
    "Rol",
    "Tipo de Personal",
    "Estado",
]

ESTADO_ACTIVO = "Activo"
ESTADO_BAJA = "Dado de baja"
ESTADOS_VALIDOS = {ESTADO_ACTIVO, ESTADO_BAJA}

MAX_ROWS = 5000

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

_SAMPLE_ROWS = (
    ("jperez", "Juan", "Pérez", "López", "juan.perez@example.com", "12345", "Médico", "Médico", "Activo"),
    ("mgomez", "María", "Gómez", "", "", "", "Recepción", "", "Activo"),
)


class ImportFileError(Exception):
    """Error de nivel-archivo (headers, extensión, límite de filas): 400, ninguna fila se evalúa."""

    def __init__(self, code: str, message: str, details: Optional[dict] = None):
        self.code = code
        self.message = message
        self.details = details or {}
        super().__init__(message)


def build_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    BRAND = "D94300"
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=BRAND)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for row_idx, sample_row in enumerate(_SAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(sample_row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, header in enumerate(HEADERS, start=1):
        from openpyxl.utils import get_column_letter

        max_len = max(len(header), *(len(str(row[col_idx - 1])) for row in _SAMPLE_ROWS))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, 12), 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _clean_text(value) -> str:
    text = "" if value is None else str(value)
    if text.strip().lower() == "nan":
        text = ""
    return re.sub(r"\s+", " ", text.strip())


def _strip_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(c for c in normalized if not unicodedata.combining(c))


def _read_dataframe(file) -> pd.DataFrame:
    filename = (getattr(file, "name", "") or "").lower()
    if not filename.endswith((".xlsx", ".xls")):
        raise ImportFileError(
            code="IMPORT_FILE_INVALID",
            message="El archivo debe ser .xlsx o .xls.",
        )

    try:
        df = pd.read_excel(file, dtype=str, engine="openpyxl" if filename.endswith(".xlsx") else None)
    except Exception as exc:
        raise ImportFileError(
            code="IMPORT_FILE_INVALID",
            message="No se pudo leer el archivo. Verifique que sea un Excel válido.",
        ) from exc

    df = df.fillna("")

    actual_headers = [str(c).strip() for c in df.columns.tolist()]
    if actual_headers != HEADERS:
        raise ImportFileError(
            code="IMPORT_HEADERS_MISMATCH",
            message="Las columnas del archivo no coinciden con la plantilla de usuarios.",
            details={"expected": HEADERS, "actual": actual_headers},
        )

    if len(df) > MAX_ROWS:
        raise ImportFileError(
            code="IMPORT_TOO_MANY_ROWS",
            message=f"El archivo excede el máximo de {MAX_ROWS} filas.",
        )

    return df


def parse_and_validate(file) -> dict:
    """
    Única función de validación, reusada tal cual por preview y por confirm --
    el confirm NUNCA confía en el resultado que mando el cliente en preview,
    siempre re-corre esto contra el archivo re-subido.

    Devuelve:
        {
            "total_records": int,
            "total_errores": int,
            "rows": [
                {
                    "row": int,             # numero de fila en el Excel (1-based, incluye header)
                    "data": {...},          # datos parseados en camelCase
                    "errors": [str, ...],   # vacio si la fila es valida
                },
                ...
            ],
        }
    """
    df = _read_dataframe(file)
    total_records = len(df)

    usernames_seen: dict = {}
    emails_seen: dict = {}

    raw_rows = []
    for row_idx in range(total_records):
        raw_rows.append(
            {
                "username": _clean_text(df.iloc[row_idx]["Usuario"]),
                "firstName": _clean_text(df.iloc[row_idx]["Nombre(s)"]),
                "paternalName": _clean_text(df.iloc[row_idx]["Apellido Paterno"]),
                "maternalName": _clean_text(df.iloc[row_idx]["Apellido Materno"]),
                "email": _clean_text(df.iloc[row_idx]["Correo"]),
                "noExp": _clean_text(df.iloc[row_idx]["No. Expediente SERMED"]),
                "roleName": _clean_text(df.iloc[row_idx]["Rol"]),
                "tipoPersonalName": _clean_text(df.iloc[row_idx]["Tipo de Personal"]),
                "estado": _clean_text(df.iloc[row_idx]["Estado"]),
            }
        )

    for parsed in raw_rows:
        if parsed["username"]:
            key = parsed["username"]
            usernames_seen[key] = usernames_seen.get(key, 0) + 1
        if parsed["email"]:
            key = parsed["email"].lower()
            emails_seen[key] = emails_seen.get(key, 0) + 1

    existing_usernames = set(
        SyUsuario.objects.filter(
            usuario__in=[p["username"] for p in raw_rows if p["username"]]
        ).values_list("usuario", flat=True)
    )
    existing_emails = set(
        email.lower()
        for email in SyUsuario.objects.filter(
            correo__in=[p["email"] for p in raw_rows if p["email"]]
        ).values_list("correo", flat=True)
        if email
    )
    role_names = {p["roleName"] for p in raw_rows if p["roleName"]}
    roles_by_name = {
        role.rol: role
        for role in Roles.objects.filter(rol__in=role_names, is_active=True)
    }
    tipo_personal_names = {p["tipoPersonalName"] for p in raw_rows if p["tipoPersonalName"]}
    tipo_personal_by_name = {
        tipo.name: tipo
        for tipo in CatTipoPersonal.objects.filter(
            name__in=tipo_personal_names, is_active=True
        )
    }

    result_rows = []
    total_errores = 0

    for row_idx, parsed in enumerate(raw_rows):
        errors = []

        if not parsed["username"]:
            errors.append("Usuario es obligatorio.")
        elif usernames_seen.get(parsed["username"], 0) > 1:
            errors.append("Usuario duplicado en el archivo.")
        elif parsed["username"] in existing_usernames:
            errors.append("Ya existe un usuario con ese nombre de usuario.")

        if not parsed["firstName"]:
            errors.append("Nombre(s) es obligatorio.")

        if parsed["email"]:
            if not _EMAIL_RE.match(parsed["email"]):
                errors.append("Correo con formato inválido.")
            else:
                email_key = parsed["email"].lower()
                if emails_seen.get(email_key, 0) > 1:
                    errors.append("Correo duplicado en el archivo.")
                elif email_key in existing_emails:
                    errors.append("Ya existe un usuario con ese correo.")

        role = None
        if not parsed["roleName"]:
            errors.append("Rol es obligatorio.")
        else:
            role = roles_by_name.get(parsed["roleName"])
            if role is None:
                errors.append(
                    f"Rol '{parsed['roleName']}' no existe o no está activo."
                )

        tipo_personal = None
        if parsed["tipoPersonalName"]:
            tipo_personal = tipo_personal_by_name.get(parsed["tipoPersonalName"])
            if tipo_personal is None:
                errors.append(
                    f"Tipo de Personal '{parsed['tipoPersonalName']}' no encontrado."
                )

        estado_normalized = parsed["estado"] or ESTADO_ACTIVO
        estado_lookup = {
            _strip_accents(v).lower(): v for v in ESTADOS_VALIDOS
        }
        estado_lookup["baja"] = ESTADO_BAJA
        estado_key = _strip_accents(estado_normalized).lower()
        if estado_key not in estado_lookup:
            errors.append("Estado debe ser 'Activo' o 'Dado de baja'.")
            estado_resolved = ESTADO_ACTIVO
        else:
            estado_resolved = estado_lookup[estado_key]

        row_data = {
            "username": parsed["username"],
            "firstName": parsed["firstName"],
            "paternalName": parsed["paternalName"],
            "maternalName": parsed["maternalName"],
            "email": parsed["email"] or None,
            "noExp": parsed["noExp"] or None,
            "roleName": parsed["roleName"],
            "roleId": role.id_rol if role else None,
            "tipoPersonalName": parsed["tipoPersonalName"],
            "tipoPersonalId": tipo_personal.id if tipo_personal else None,
            "estado": estado_resolved,
            "isActive": estado_resolved == ESTADO_ACTIVO,
        }

        result_rows.append(
            {
                "row": row_idx + 2,  # +1 header, +1 para 1-based
                "data": row_data,
                "errors": errors,
            }
        )
        if errors:
            total_errores += 1

    return {
        "total_records": total_records,
        "total_errores": total_errores,
        "rows": result_rows,
    }
